# classifier.py
# -*- coding: utf-8 -*-

import re
import os
from pathlib import Path
from typing import Dict, List, Tuple
import pandas as pd

try:
    from tqdm import tqdm
    tqdm.pandas(desc="  Classifying")
    USE_TQDM = True
except ImportError:
    USE_TQDM = False

def _candidate_bases() -> list[Path]:
    bases: list[Path] = []
    user = Path(os.environ.get("USERPROFILE", ""))
    p_base = user / "PETRONAS"
    if p_base.exists():
        bases.append(p_base)
    for envvar in ("OneDriveCommercial", "OneDrive"):
        v = os.environ.get(envvar)
        if v and os.path.isdir(v):
            bases.append(Path(v))
    od_guess = user / "OneDrive - PETRONAS"
    if od_guess.exists():
        bases.append(od_guess)
    seen, uniq = set(), []
    for b in bases:
        s = str(b)
        if s not in seen:
            seen.add(s); uniq.append(b)
    return uniq

BASES = _candidate_bases()

def _first_existing(*rel_parts: str) -> Path | None:
    for b in BASES:
        p = b.joinpath(*rel_parts)
        if p.exists():
            return p
    return None

_inc_env = os.environ.get("INCIDENT_ROOT")
if _inc_env:
    INCIDENT_ROOT = Path(_inc_env)
else:
    INCIDENT_ROOT = _first_existing("TRMS Internal - myGenie+ Extract")
    if INCIDENT_ROOT is None:
        INCIDENT_ROOT = Path(os.environ["USERPROFILE"]) / r"PETRONAS\TRMS Internal - myGenie+ Extract"

OUT_DIR = (
    Path(os.environ["USERPROFILE"])
    / "PETRONAS"
    / "PETCO Trading Digital - myGenie Ticket Analysis"
    / "Ticket Categorization"
)
OUT_DIR.mkdir(parents=True, exist_ok=True)
OUTPUT_FILE = OUT_DIR / "Classified_Incidents.xlsx"

print("Resolved paths:")
print(f"  INCIDENT_ROOT : {INCIDENT_ROOT}")
print(f"  OUTPUT_FILE   : {OUTPUT_FILE}")

if not INCIDENT_ROOT.exists():
    raise FileNotFoundError(f"INCIDENT_ROOT not found: {INCIDENT_ROOT}")

def find_latest_incident_file(folder: Path) -> Path:
    pattern = re.compile(r"^Incident Raw Data - (\d{1,2}) (\w+) (\d{4})\.xlsx$", re.IGNORECASE)
    candidates = []
    for f in folder.iterdir():
        if not f.is_file():
            continue
        m = pattern.match(f.name)
        if m:
            try:
                dt = pd.to_datetime(f"{m.group(1)} {m.group(2)} {m.group(3)}", dayfirst=True)
                candidates.append((dt, f))
            except Exception:
                continue
    if not candidates:
        raise FileNotFoundError(f"No 'Incident Raw Data - <dd> <Mon> <yyyy>.xlsx' found in:\n  {folder}")
    candidates.sort(key=lambda x: x[0], reverse=True)
    latest_dt, latest_file = candidates[0]
    print(f"  Selected: {latest_file.name}  (date: {latest_dt.date()})")
    return latest_file

INPUT_FILE = find_latest_incident_file(INCIDENT_ROOT)

COL_DETAILED_DESC_1 = "Detailed Decription"
COL_DETAILED_DESC_2 = "Detailed Description"
COL_PCT3            = "Product Categorization Tier 3"
COL_DESC_EXTRACT    = "Desc Extract"
ID_COLS_PREF        = ["Service Request ID", "Incident ID"]
DATE_COLS_PREF      = [
    "Last Modified Date", "Actual Resolution Date", "Closed Date",
    "Last Resolved Date", "Reported Date", "Actual Reported Date", "Re-Opened Date"
]

def norm_whitespace(x: str) -> str:
    x = x.replace("\u3000", " ")
    x = x.replace("\r\n", "\n").replace("\r", "\n")
    x = re.sub(r"[ \t]+", " ", x)
    x = re.sub(r"\n[ \t]+", "\n", x)
    return x.strip()

def to_datetime_smart(s: pd.Series) -> pd.Series:
    if pd.api.types.is_datetime64_any_dtype(s):
        return s
    parsed = pd.to_datetime(s, errors="coerce")
    s_num  = pd.to_numeric(s, errors="coerce")
    serial_mask = s_num.between(1, 80000)
    if serial_mask.any():
        try:
            parsed_serial = pd.to_datetime(s_num, errors="coerce", unit="D", origin="1899-12-30")
        except Exception:
            parsed_serial = pd.Series(pd.NaT, index=s.index)
    else:
        parsed_serial = pd.Series(pd.NaT, index=s.index)
    return parsed if parsed.notna().sum() >= parsed_serial.notna().sum() else parsed_serial

DESC_START_COMPILED = [re.compile(p, re.IGNORECASE | re.DOTALL) for p in [
    r"(?is)description\s*[::_-]\s*(.*?)\bplease\s*provide\s*[::_-]",
    r"(?is)description\s*[::_-]\s*(.*)$",
    r"(?is)\bdesc\s*[::_-]\s*(.*?)\bplease\s*provide\s*[::_-]",
    r"(?is)\bdesc\s*[::_-]\s*(.*)$",
    r"(?is)\bissue\s*[::_-]\s*(.*?)\bplease\s*provide\s*[::_-]",
    r"(?is)\bissue\s*[::_-]\s*(.*)$",
    r"(?is)^issue\s*[-]\s*(.+)$",
]]

def extract_desc_only(cell: str) -> str:
    if not isinstance(cell, str) or not cell.strip():
        return ""
    cell = norm_whitespace(cell)
    for rx in DESC_START_COMPILED:
        m = rx.search(cell)
        if m:
            snippet = norm_whitespace(m.group(1))
            if snippet:
                return snippet
    lines = cell.split("\n")
    desc_started = False
    chunks: List[str] = []
    for ln in lines:
        ln_clean = ln.strip()
        if re.match(r"(?i)^description\s*[::_-]?\s*$", ln_clean):
            desc_started = True
            continue
        if not desc_started:
            m2 = re.match(r"(?is)^description\s*[::_-]\s*(.+)$", ln_clean)
            if m2:
                desc_started = True
                chunks.append(norm_whitespace(m2.group(1)))
            continue
        if re.match(r"(?i)^please\s*provide\s*[::_-]?\s*$", ln_clean) or \
           re.search(r"(?i)^please\s*provide\s*[::_-]", ln_clean):
            break
        chunks.append(ln_clean)
    snippet2 = norm_whitespace(" ".join([c for c in chunks if c]))
    return snippet2 if snippet2 else norm_whitespace(cell)

VERBS_AMEND = r"(amend|fix|correct|update|change|edit|adjust|revise|modify|add|remove|delete)"
VERBS_PUSH  = r"(push|trigger|retrigger|re-push|re push|send to)"
VERBS_POST  = r"(post|posting|posted)"
AMOUNT_TERMS = r"(amount|myr|usd|local\s*currency|lc\s*[0-9]+)"
PRICE_TERMS = (
    r"(price|unit\s*price|pricing|published\s*price|wma|jkm|bpp|"
    r"strike\s*price|settlement\s*price|forward\s*curve|forward\s*price)"
)
PRICE_TERMS_BROAD = (
    r"(price|unit\s*price|pricing|published\s*price|wma|jkm|bpp|"
    r"strike\s*price|mtm|p&l|pnl|pnl_cy|pnl_ly|"
    r"eod|settlement\s*price|forward\s*curve|forward\s*price|ice|nymex|henry\s*hub|ttf|nbp|jcc)"
)
FX_TERMS = (
    r"(exchange\s*rate|fx\s*rates?|fx\s*rate|fx|"
    r"currency\s*conversion|eur\s*to\s*usd|usd\s*to\s*myr|myr\s*conversion|"
    r"fx\s*rates?\s*undefined|not\s*recognis(e|z)ed?\s*fx)"
)
DATE_TERMS = (
    r"(posting\s*date|invoice\s*date|accounting\s*date|period|backdate|date|"
    r"bl\s*date|movement\s*date|supplier(s'?)*\s*invoice\s*date|period\s*1)"
)
TAX_TERMS  = r"(tax|vat|gst|withholding|sst)"
DOC_TERMS  = r"(invoice|inv\.?|document|doc(ument)?\s*(no|number)?)"
QTY_TERMS  = r"(qty|quantity|volume|bbls?|barrels?|mt|tons?|mmbtu|units?|m3|kt)"
BANK_TERMS = r"(bank\s*account|iban|swift|beneficiary\s*account|acct\s*no|account\s*number|account\s*no)"
MISSING_PHRASES = (
    r"(missing|not\s*(appear(ing)?|shown|sighted|populated|reflect(ed)?|flow(ing)?|found|available|visible)|"
    r"no\s*(data|vendor\s*code)\s*(appear|shown)|did\s*not\s*populate|does\s*not\s*appear|did\s*not\s*appear|"
    r"not\s*auto.?populate(d)?|not\s*pick(ed)?\s*up|not\s*include(d)?)"
)
RUN_CORR_PHRASES = r"(run\s*corrective|corrective\s*run|please\s*run\s*corrective|fsa\s*0)"
JE_NE_SAP        = r"(je|new\s*entries|ne|sap|journal\s*entries?)(\s*tab)?"
SYSTEM_ERROR_PHRASES = (
    r"(error|failed|fail(ure)?|blocked|unsuccessful|cannot|unable to|can't|pop.?up|pops?\s*out|"
    r"exception|unresponsive|not\s*responding|access\s*denied|security\s*enforced|"
    r"nomination\s*(save\s*)?failed|save\s*failed|actuali[sz]ation\s*error|system\s*issue|"
    r"bug|glitch|citrix|login|log\s*in|access\s*issue)"
)
DEAL_DUP_SIGNALS = (
    r"(duplicate\s*(deal|invoice|ifp|doc(ument)?)|"
    r"duplication\s*of\s*deals?|"
    r"same\s*reference\s*number|"
    r"double\s*posting\s*(in\s*)?sap|"
    r"appear(s|ed)?\s*(twice|double)|"
    r"(invoice|deal)\s*(appear|appear\s*twice|duplicated))"
)

CATS = [
    "Discrepancies - Customer Code", "Discrepancies - Company Code", "Discrepancies - Legal Entity",
    "Discrepancies - Account", "Discrepancies - Broker broker", "Discrepancies - Strategy",
    "Discrepancies - Date", "Discrepancies - Amount", "Discrepancies - Tax", "Discrepancies - Doc",
    "Discrepancies - Address", "Discrepancies - exchange rate", "Discrepancies - vessel",
    "Discrepancies - Bank Account Issue", "Discrepancies - Details", "Discrepancies - Counterparty",
    "Discrepancies - Freight", "Discrepancies - Bunker", "run corrective", "Missing Data",
    "Duplicate Invoice/Deal Number", "SAP Posting Failures", "Others", "Question Help Support",
    "Not Update Data", "Push to SAP", "Push Document", "Wrong Price", "Wrong Quantity",
    "System Error / Functional Issue", "Report Issue", "Access Issue",
    "PLSB Bucket Amendment", "Matching Issue", "New Setup / Configuration",
]

PATTERNS: Dict[str, List[Tuple[str, int, str]]] = {

    "run corrective": [
        (r"\b" + RUN_CORR_PHRASES + r"\b", 12, "run corrective keyword"),
        (r"\bfsa\s*\d{3}\b", 10, "FSA error code"),
        (r"(no\s*authorization|period\s*0\d{2}\s*\d{4}|not\s*authorized.*period)", 9, "period/authorization error"),
        (r"\b(corrective|fix)\b.{0,40}\b(double|duplicate)\s*(line|entry)\b", 10, "fix double line"),
        (r"\bmissing\s*another\s*\d*\s*lines?\b", 9, "missing another lines"),
        (r"\bonly\s*(tax|adjustment|barging)\s*line\s*(appear|populate)\b", 10, "only tax/adj/barging line appeared"),
        (r"\bmissing\s*(gl|cost|principal)\s*line\b", 10, "missing GL/cost/principal line"),
        (r"\bonly\s*(appeared|appear)\s*.{0,40}\bi[/\\]o\b", 11, "only appeared X i/o Y"),
        (r"\bi[/\\]o\s*(barging|cargo|product|cost|secondary)", 11, "i/o line item pattern"),
        (r"\bduplicate\s*line\s*items?\b", 12, "duplicate line items"),
        (r"\bdouble\s*line\s*items?\b", 12, "double line items"),
        (r"\b(duplicate|double)\s*(barging|product\s*cost|secondary\s*cost|cargo\s*line)\b", 12, "duplicate barging/product cost"),
        (r"\btriplet\s*line\b|\btriple\s*line\b", 11, "triplet line"),
        (r"\bexcess\s*(line|item)\b.{0,20}\b(ne|new\s*entries|doc)\b", 9, "excess line in NE"),
        (r"\bremove\s*(the\s*)?(extra|duplicate|double)\s*(line|item)\b", 10, "remove extra line"),
        (r"\bonly\s*(appeared|appear)\s*(barging|cargo|product|secondary|cost)\b.{0,30}\bi[/\\]o\b", 12, "only appeared X i/o"),
        (r"\bno\s*(product\s*(cost|sales)|cargo\s*line|sales\s*line)\s*(appear|populated|shown)\b", 10, "no product cost/sales appeared"),
        (r"\bto\s*run\s*corr", 11, "to run corrective shortform"),
        (r"\bmissing\s*(line|gl|item)\b", 7, "missing line/gl/item"),
        (r"\b(only\s*)?\d\s*line\s*item\s*(appear|populated|come\s*out)\b.{0,30}\b(should\s*be|i[/\\]o|instead)\b", 9, "N line items wrong count"),
    ],

    "Duplicate Invoice/Deal Number": [
        (r"\bduplicate\b.{0,30}\b(invoice|deal)\b", 13, "duplicate invoice/deal"),
        (r"\b(invoice|deal)\b.{0,30}\bduplicate\b", 13, "invoice/deal duplicate"),
        (r"\bduplication\s*of\s*deals?\b", 13, "duplication of deals"),
        (r"\bduplicate\s*deals?\s*in\s*endur\b", 13, "duplicate deals in endur"),
        (r"\bremove\s*duplicate\b.{0,20}\b(deal|invoice)\b", 12, "remove duplicate deal/invoice"),
        (r"\bexecuted\s*\d+\s*but\s*(in\s*endur|system)\s*\d+\b", 12, "executed vs system mismatch"),
        (r"\bappear(s|ed)?\s*(twice|double)\b.{0,30}\b(deal|invoice|cargo)\b", 12, "deal appears twice"),
        (r"\bsame\s*reference\s*number\b", 11, "same reference number"),
        (r"\bdouble\s*posting\s*(in\s*)?sap\b", 12, "double posting in SAP"),
        (r"\bduplication\s*of\s*paper\s*trade\b", 12, "duplication of paper trade"),
        (r"\bduplicate(d)?\s*cargo\s*(id|ids?)\b", 11, "duplicated cargo ID"),
        (r"\bdelete\s*(this\s*)?(claim\s*id|cl-\d+)\b.{0,20}\bdue\s*to\s*duplicate\b", 11, "delete claim ID duplicate"),
        (r"\bduplicate\b.{0,20}\b(doc(ument)?|ifp)\b", 9, "duplicate document/IFP"),
        (r"\b(duplicate|duplication)\s*(entr(y|ies)|item)\b", 8, "duplicate entries generic"),
        (r"\bappear(s|ed)?\s*(twice|double|duplicate)\b", 8, "appears twice/double"),
        (r"\bremove\s*duplicate\b", 9, "remove duplicate"),
        (r"\bduplicates?\s*(item|items?|line)\s*(in\s*endur|endur\s*system|system)\b", 11, "duplicates item in endur"),
    ],

    "Push Document": [
        # Pushing documents to JE (Journal Entries) or NE (New Entries) tabs
        # These are document staging actions — NOT direct SAP posting
        (r"\b" + VERBS_PUSH + r"\b.{0,40}\b(ne|new\s*entries|je|journal\s*entries?)(\s*tab)?\b", 12, "push to NE/JE tab"),
        (r"\bpush.{0,20}\bne\b|\bne.{0,20}\bpush\b|\bpush.{0,20}\bje\b|\bje.{0,20}\bpush\b", 11, "push NE/JE"),
        (r"\binto\s*(ne|je|new\s*entries|journal\s*entries)\s*tab\b", 11, "push into NE/JE tab"),
        (r"\bpush\s*(into|to)\s*(ne|je|new\s*entries|journal\s*entries)\b", 11, "push into NE/JE explicit"),
        (r"\bpush\s*(doc\s*)?(no\s*)?\d+\s*to\s*(ne|je|new\s*entries)\b", 13, "push doc to NE/JE"),
        (r"\bpush\s*doc(ument)?\s*(no|number)?\s*\d+", 11, "push doc number"),
        (r"\bpush\s*doc\b", 11, "push doc"),
        (r"\bpush\s*to\s*drafted\b", 11, "push to drafted"),
        (r"\bhelp\s*to\s*push\b", 10, "help to push"),
        (r"\bplease\s*(help\s*to\s*)?push\b", 10, "please push"),
        (r"\bpush.{0,20}\btab\b", 8, "push tab"),
        (r"\burgently?.{0,20}\bpush\b|\bpush\b.{0,20}\burgently?\b", 10, "urgently push"),
        (r"\bnot\s*(appear|populate)\b.{0,30}\b(ne|new\s*entries)\b", 8, "not appear in NE"),
        (r"\b(ne|new\s*entries)\b.{0,30}\bnot\s*(appear|populate)\b", 8, "NE not appear"),
        (r"\bpull\s*back\b.{0,20}\b(ne|new\s*entries|doc|document)\b", 10, "pull back to NE"),
        (r"\btransfer\s*doc\b.{0,30}\b(ne|from|month|november|december|january)\b", 9, "transfer doc NE month"),
        (r"\bmove\s*doc\b.{0,30}\b(ne|month|status)\b", 9, "move doc NE"),
    ],

    "Push to SAP": [
        # Pushing directly to SAP — the actual SAP system posting action
        (r"\bpush\s*(doc\s*)?(no\s*)?\d+\s*to\s*sap\b", 13, "push doc to SAP explicitly"),
        (r"\b" + VERBS_PUSH + r"\b.{0,40}\bsap\b", 12, "push to SAP keyword"),
        (r"\bkindly\s*(help\s*to\s*)?push\b.{0,30}\bsap\b", 11, "kindly push to SAP"),
        (r"\bplease\s*(help\s*to\s*)?push\b.{0,30}\bsap\b", 11, "please push to SAP"),
        (r"\bpush\s*to\s*sap\b", 13, "push to SAP direct"),
        (r"\bsend\s*to\s*sap\b|\btrigger\s*to\s*sap\b|\bre.?push\s*to\s*sap\b", 12, "send/trigger/repush to SAP"),
    ],

    "SAP Posting Failures": [
        (r"\b" + VERBS_POST + r"\b.{0,30}\b(fail(ed)?|unsuccessful|error|blocked)\b", 11, "posting failed/blocked"),
        (r"\bblocked\s*for\s*posting\b", 12, "blocked for posting"),
        (r"\bje\s*failed\b|\bposting\s*failed\b", 11, "JE/posting failed"),
        (r"\b(sap\s*doc|posting\s*number)\b.{0,40}\b(not\s*appear|hasn.?t\s*appear|did\s*not\s*appear)\b", 11, "SAP posting not appeared"),
        (r"\bfailed\s*(to\s*generate|entries)\b.{0,30}\bflatfile\b", 9, "failed to generate flatfile"),
        (r"\b(rw\s*\d{3}|no\s*item\s*information\s*transferred)\b", 10, "SAP transfer error code"),
        (r"\bblocked\s*account\s*for\s*posting\b|\baccount.{0,20}\bblocked\b", 10, "account blocked posting"),
        (r"\b(sap\s*posting|sap\s*doc(ument)?)\b.{0,30}\b(not\s*(appear|generated)|missing)\b", 10, "SAP doc not appearing"),
        (r"\bfailed\s*(entries|entry)\b", 9, "failed entries"),
        (r"\bedoc(ument)?\s*(for\s*invoice\s*)?not\s*found\b", 11, "edoc not found"),
        (r"\bbalance\s*in\s*(transaction|local)\s*currency\b", 11, "balance in transaction currency"),
        (r"\b(f5\s*703|rw\s*033|zval_message)\b", 11, "SAP error code F5/RW"),
        (r"\bduplicate\s*source\s*system\s*trans(action)?\s*id\b", 11, "duplicate source system trans ID"),
        (r"\binvalid\s*input\b.{0,30}\b(account|profit\s*center|trading\s*partner)\b", 10, "invalid input SAP field"),
        (r"\b(not\s*appear|does\s*not\s*appear|not\s*posted)\b.{0,20}\b(sap)\b", 10, "not posted in SAP"),
        (r"\bsap\b.{0,20}\b(not\s*appear|not\s*posted|missing)\b", 10, "SAP not appear"),
        (r"\bsystem\s*(closed|auto\s*close)\b.{0,30}\b(posting|push|sap)\b", 9, "system closed during posting"),
        (r"\b(resubmit|re.?submit)\b.{0,30}\b(je|sap|posting|flatfile)\b", 9, "resubmit JE/SAP"),
        (r"\b(successfully\s*posted|posted\s*to\s*sap)\b.{0,60}\b(reappear|appear\s*again|still\s*appear|appear\s*back)\b", 11, "doc reappeared in NE after SAP"),
        (r"\b(reappear|appear\s*again|appear\s*back)\b.{0,40}\b(ne|new\s*entries|sap)\b", 10, "reappeared in NE"),
        (r"\bno\s*sap\s*reference\s*doc\s*(no|number)\b", 10, "no SAP ref doc no"),
        (r"\bkindly\s*post\s*to\s*sap\b|\bplease\s*post\s*(to\s*)?sap\b", 10, "kindly/please post to SAP"),
        (r"\b(in\s*progress|still\s*(appear|shows?)\s*in\s*progress)\b.{0,30}\b(ne|new\s*entries)\b", 9, "stuck in progress NE"),
    ],

    "Not Update Data": [
        (MISSING_PHRASES + r".{0,30}\b(ne|new\s*entries|je|journal)\b", 10, "not appearing in NE/JE"),
        (r"(selling\s*price|price)\s*(not\s*appear(ed)?|not\s*updated?|not\s*reflect(ed)?)", 9, "price not appeared/updated"),
        (r"\bnot\s*updated?\b", 8, "not updated"),
        (r"\bnot\s*appear(ing)?\b", 8, "not appearing"),
        (r"\bnot\s*sighted\b|\bnot\s*sighted\s*in\b", 8, "not sighted"),
        (r"\b(price|pricing)\s*not\s*(reflect(ed)?|shown|updated?|appear(ing)?)\b", 9, "price not reflected/shown"),
        (r"\b(selling\s*price|physical\s*pricing)\s*not\s*appear(ed)?\b", 10, "selling price not appeared"),
        (r"\b(issue|incorrect)\s*soa\s*type\b", 8, "incorrect SOA type"),
        (r"\b(price|pricing)\b.{0,30}\b(still\s*)?(unknown|not\s*known)\b", 11, "pricing still unknown"),
        (r"\b(deal|cargo|item)\s*(not\s*appear|did\s*not\s*appear|does\s*not\s*appear)\b", 8, "deal/cargo not appearing"),
        (r"\bdoes\s*not\s*reflect\b|\bnot\s*reflect(ed|ing)?\b", 8, "not reflecting"),
        (r"\bstuck\s*in\s*(sent\s*to\s*lhdn|status)\b", 9, "stuck in LHDN/status"),
        (r"\b(price|pricing)\s*still\s*(remain|unchanged|not\s*updated?)\b", 9, "price not updated"),
        (r"\bnot\s*flow(ing)?\s*(through|to|in)\b", 8, "not flowing through"),
        (r"\bnot\s*appear(ing)?\s*in\s*(receivable|payable)\s*standard\b", 11, "not in receivable/payable standard"),
        (r"\b(delivery|deal)\s*id\b.{0,30}\b(not\s*appear|did\s*not\s*appear|does\s*not\s*appear)\b", 10, "delivery/deal id not appearing"),
        (r"\b(do(?:es)?\s*not\s*appear|missing)\b.{0,30}\b(physical\s*invoic(ing)?|payable\s*standard|desktop)\b", 10, "not in invoicing desktop"),
        (r"\bnot\s*pick(ed)?\s*up\b", 8, "not picked up"),
        (r"\bstatus\s*(stuck|not\s*changed|still\s*shows?)\b", 8, "status stuck"),
        (r"\b(paper\s*trade|deal|trade)s?\b.{0,30}\b(not\s*(flow|appear|autoflow|auto.?flow|captured)|didn.?t\s*(flow|appear|capture))\b", 10, "trades not flowing"),
        (r"\b(auto.?capture|auto.?flow)\b.{0,30}\b(not|didn.?t|failed|issue)\b", 10, "autocapture/autoflow issue"),
        (r"\btrades?\s*(stopped|stop)\s*(auto.?flow|flow)\b", 10, "trades stopped autoflowing"),
        (r"\bdeals?\b.{0,30}\b(did\s*not\s*display|not\s*display|not\s*captured)\b.{0,30}\b(endur|system)\b", 9, "deals not display in endur"),
        (r"\b(brn|efs|dbi)\s*(leg|handle)\b.{0,30}\b(not\s*flow|did\s*not\s*flow)\b", 10, "BRN/EFS/DBI leg not flowing"),
        (r"\bstuck\s*(at|in)\s*(sent\s*to\s*lhdn|lhdn)\b", 10, "stuck at LHDN"),
        (r"\b(pending|doc\s*pending)\b.{0,20}\b(qr\s*(validation|code)|lhdn|irb)\b", 10, "pending QR LHDN/IRB"),
        (r"\bcargo\s*status\b.{0,40}\b(not\s*change|did\s*not\s*change|still)\b.{0,20}\b(imos\s*received|imos)\b", 10, "cargo status not IMOS received"),
        (r"\bsend\s*to\s*imos\s*confirm(ed)?\b.{0,30}\b(not\s*change|stuck|still)\b", 10, "send to IMOS confirmed stuck"),
        (r"\b(move|help\s*to\s*move)\b.{0,20}\b(doc|document)\b.{0,20}\b(status\s*to\s*approved|approved)\b", 9, "move doc to approved"),
        (r"\bneed\s*to\s*move\b.{0,20}\bdoc\b.{0,20}\bapproved\b", 9, "need to move doc to approved"),
    ],

    "Missing Data": [
        (r"\bmissing\s*tax\s*line\b", 12, "missing tax line"),
        (r"\bmissing\s*line\s*(item|gl)?\b", 10, "missing line item/gl"),
        (r"\bmissing\b", 5, "generic missing"),
        (r"\bmain\s*line(s)?\s*missing\b", 10, "main line missing"),
        (r"\b(tax\s*code|tax\s*type)\s*(did\s*not\s*reflect|not\s*reflect(ed)?|not\s*appear(ing)?|did\s*not\s*appear|does\s*not\s*appear)\b", 11, "tax code not reflected"),
        (r"\btax\s*code\s*did\s*not\s*reflect\b", 12, "tax code did not reflect"),
        (r"\btax\s*(code|type)\s*(not\s*auto.?pop|not\s*populat|not\s*found|configuration\s*not\s*found)\b", 11, "tax code config not found"),
        (r"\bno\s*tax\s*(type|code|line)\b", 10, "no tax type/code/line"),
        (r"\b(vendor\s*code|sap\s*doc\s*type)\s*(not\s*auto.?pop|did\s*not\s*pop|not\s*reflect|does\s*not\s*appear|missing)\b", 10, "vendor code/SAP doc type missing"),
        (r"\bgl\s*(account)?\s*(disappear|missing|not\s*appear|fee\s*holding)\b", 10, "GL account missing"),
        (r"\b(cost|item|fee|adjustment)\s*line\s*(missing|not\s*appear|did\s*not\s*appear)\b", 9, "cost/item line missing"),
        (r"\bmissing\s*(cost|item|fee)\s*line\b", 10, "missing cost/item/fee line"),
        (r"\bno\s*vendor\s*code\b", 9, "no vendor code"),
        (r"\bno\s*sap\s*doc\s*type\b|\bsap\s*doc\s*type.{0,20}missing\b", 10, "SAP doc type missing"),
        (r"\bdelivery\s*id\s*(missing|not\s*appear|unable\s*to\s*(upload|find))\b", 9, "delivery ID missing"),
        (r"\b(no\s*)?(tax\s*(code|type|line)|vendor\s*code|sap\s*doc\s*type)\s*(not\s*(appear|reflect|populate|auto.?pop)|missing|did\s*not|does\s*not)\b", 12, "structural config missing"),
        (r"\btax\s*(code|type)\s*not\s*(auto.?pop|populat)\b", 11, "tax not auto-populated"),
        (r"\bno\s*tax\s*line\b.{0,30}\b(desktop|invoicing|physical)\b", 11, "no tax line in desktop"),
        (r"\b(reference\s*key|cost\s*cent(re|er)|profit\s*cent(re|er))\s*(not\s*appear|missing|blank)\b", 9, "reference key/cost centre missing"),
        (r"\b(register|add)\s*(vendor\s*code|counterpart)\b.{0,30}\b(endur|backend|system|trms)\b", 10, "register vendor code endur"),
        (r"\bvendor\s*code\b.{0,30}\b(register|add|insert|not\s*in\s*endur|backend)\b", 10, "vendor code not in endur"),
        (r"\bfee\s*type\b.{0,30}\b(not\s*appear|missing)\b.{0,20}\b(ifp|invoice)\b", 9, "fee type not in IFP"),
        (r"\bno\s*(return|result|line\s*item)\b.{0,20}\b(query|desktop|endur)\b", 8, "no return in query"),
        (r"\bnothing\s*appears?\b.{0,20}\b(query|desktop|endur|je|ne)\b", 8, "nothing appears in query"),
        (r"\bno\s*(dropdown|option|drop.?down)\b.{0,20}\b(remittance|bank|settle|move\s*to\s*status)\b", 9, "no dropdown option"),
        (r"\b(payable|receivable)\b.{0,20}\b(tab|desktop|window)\b.{0,30}\bno\s*(transaction|data|item)\s*(appear|show)\b", 9, "payable tab no transactions"),
        (r"\b(product|instrument|ticker)\b.{0,20}\bnot\s*(in\s*)?(the\s*)?(product\s*list|system|endur|list|injection\s*option)\b", 8, "product not in list"),
        (r"\bbroker\s*(fee)?\b.{0,30}\bnot\s*(includ|captur|coming\s*into|flow)\b.{0,20}\b(auto.?capture|endur|deal)\b", 9, "broker fee not in autocapture"),
    ],

    "Wrong Price": [
        # REQUIRE both a problem word AND a price-related term — prevents over-classification
        (r"\b(wrong|incorrect|not\s*(same|match(ed)?|matching)|mismatch(ed)?|misalign(ed)?)\b.{0,60}\b" + PRICE_TERMS + r"\b", 11, "wrong/incorrect price"),
        (r"\b" + PRICE_TERMS + r"\b.{0,60}\b(wrong|incorrect|not\s*same|mismatch(ed)?)\b", 11, "price wrong/different"),
        # Broad price terms also require a problem word
        (r"\b(wrong|incorrect)\b.{0,40}\b" + PRICE_TERMS_BROAD + r"\b", 10, "wrong broad price term"),
        (r"\b" + PRICE_TERMS_BROAD + r"\b.{0,40}\b(wrong|incorrect)\b", 10, "broad price term wrong"),
        # Explicit category labels — high confidence
        (r"\bcategory\s*of\s*issue\s*[:]\s*wrong\s*price\b", 13, "category wrong price"),
        (r"\bwrong\s*strike\s*price\b", 13, "wrong strike price"),
        (r"\bcategory\s*of\s*issue\s*[:]\s*wrong\s*final\s*settlement\s*price\b", 13, "wrong final settlement price"),
        (r"\b(strike\s*price|contract\s*price)\s*(in\s*endur|endur)\s*:?\s*\d+", 11, "strike price in endur"),
        # MTM/PnL require explicit wrong/incorrect
        (r"\b(wrong|incorrect)\s*(mtm|p&l|pnl)\s*(price|value)\b", 11, "wrong MTM/PnL value"),
        (r"\b(incorrect|wrong)\s*(eod|settlement)\s*price\b", 11, "incorrect EOD/settlement price"),
        # Decimal place issues on price fields
        (r"\b(decimal\s*place|3\s*dp|3dp)\b.{0,30}\b(price|strike|pv|fee)\b", 9, "decimal place issue on price"),
        # Price discrepancy — requires both "price" and "discrepancy/not tally"
        (r"\bprice\b.{0,40}\b(discrepancy|not\s*tally|does\s*not\s*tally)\b", 10, "price not tally"),
        (r"\bdiscrepancy\b.{0,30}\b(price|strike)\b", 10, "discrepancy in price"),
        # System showing wrong price — explicit
        (r"\b(endur|system)\b.{0,40}\b(show(ing)?|reflect(ing)?)\b.{0,40}\b(wrong|incorrect|inaccurate)\s*(price|mtm)\b", 10, "system showing wrong price"),
        (r"\b(captured|using)\s*(holiday|public\s*holiday)\s*price\b", 11, "captured holiday price"),
        (r"\b(please\s*)?remove\s*published\s*price\b", 11, "remove published price"),
        # Forward/WMA — require wrong/incorrect
        (r"\b(forward\s*price|forward\s*curve)\b.{0,40}\b(wrong|incorrect|not\s*match|differ)\b", 9, "forward price wrong"),
        (r"\b(wma|whole\s*month\s*average)\b.{0,40}\b(wrong|incorrect|not\s*reflect|differ)\b", 10, "WMA wrong"),
        # "price different" — only if both words present in close proximity
        (r"\bprice\b.{0,20}\bdifferent\b|\bdifferent\b.{0,20}\bprice\b", 8, "price different"),
        (r"\b(incorrect|wrong)\s*(premium|index)\b", 9, "incorrect premium/index"),
        # Reset/remove price — explicit action on price field
        (r"\b(reset|remove)\b.{0,15}\b(published\s*)?price\b", 9, "reset/remove price"),
    ],

    "Wrong Quantity": [
        (r"\b(wrong|incorrect|not\s*(same|match))\b.{0,40}\b" + QTY_TERMS + r"\b", 10, "wrong quantity"),
        (r"\b" + QTY_TERMS + r"\b.{0,40}\b(wrong|incorrect|not\s*same|mismatch)\b", 10, "quantity wrong"),
        (r"\b(volume|qty)\s*(incorrect|wrong|mismatch|not\s*correct)\b", 9, "volume/qty incorrect"),
        (r"\b(bbl|barrels?)\s*(instead\s*of|i/o)\s*(mt|metric\s*ton)\b", 10, "BBL instead of MT"),
        (r"\b(mt|metric\s*ton)\s*(instead\s*of|i/o)\s*(bbl|barrels?)\b", 10, "MT instead of BBL"),
        (r"\blot\s*size\b.{0,30}\b(wrong|incorrect|should\s*be|change|1000|100)\b", 9, "lot size wrong"),
        (r"\b(1000\s*i/o\s*100|100\s*i/o\s*1000)\b", 10, "lot size 1000 i/o 100"),
        (r"\bcontract\s*size\b.{0,30}\b(wrong|incorrect|should\s*be|1000|100)\b", 9, "contract size wrong"),
        (r"\b(mmbtu|delivery\s*ticket)\b.{0,30}\b(not\s*tally|incorrect|wrong)\b", 9, "MMBTU not tally"),
    ],

    "Discrepancies - Amount": [
        (r"\bdiscrepanc(y|ies)\b.{0,30}\b" + AMOUNT_TERMS + r"\b", 11, "discrepancy amount"),
        (r"\b" + AMOUNT_TERMS + r"\b.{0,30}\b(discrepancy|not\s*tally|does\s*not\s*tally|different|tally)\b", 11, "amount not tally"),
        (r"\b(amend|fix|correct)\b.{0,20}\b(usd|myr).{0,20}amount\b", 10, "amend USD/MYR amount"),
        (r"\bdifferent\b.{0,30}\b(amount|myr|usd|lc|local\s*currency)\b", 9, "different amount/currency"),
        (r"\b(amount|myr|usd)\b.{0,30}\b(different|does\s*not\s*tally|not\s*tally|does\s*not\s*match|mismatch)\b", 10, "amount mismatch"),
        (r"\bdiscrepanc(y|ies)\s*of\s*(usd|myr)\s*[\d.,]+\b", 11, "discrepancy of USD/MYR value"),
        (r"\b(local\s*currency\s*[23]|lc\s*[23])\s*(amount)?\s*(not\s*same|different|mismatch|not\s*tally)\b", 10, "LC2/LC3 mismatch"),
        (r"\b(lc3|local\s*currency\s*3)\s*amount\s*does\s*not\s*reflect\b", 11, "LC3 does not reflect"),
    ],

    "Discrepancies - exchange rate": [
        (r"\b(exchange\s*rate|fx\s*rates?|fx)\b", 11, "FX/exchange rate"),
        (r"\bfx\s*rates?\s*undefined\b", 11, "FX rates undefined"),
        (r"\b(incorrect|wrong)\b.{0,30}\b(exchange\s*rate|fx|conversion)\b", 10, "incorrect exchange rate"),
        (r"\b(conversion|converted)\s*(myr|usd|amount)\s*(wrong|incorrect|not\s*tally|different)\b", 10, "conversion amount wrong"),
        (r"\b(wrong|incorrect)\s*(currency\s*conversion|conversion)\b", 10, "wrong conversion"),
        (r"\bbnm\s*rate\b.{0,30}\b(not\s*tally|different|wrong|incorrect)\b", 10, "BNM rate not tally"),
    ],

    "Discrepancies - Date": [
        (r"\b(posting\s*date|invoice\s*date|accounting\s*date)\b", 11, "specific date field"),
        (r"\bperiod\s*(1|block|jan|january|dec|december)\b", 10, "period change"),
        (r"\b(change|amend|correct|update)\b.{0,30}\b(date|period)\b", 9, "change date/period"),
        (r"\bbackdate\b", 9, "backdate"),
        (r"\b(update|amend|correct)\b.{0,30}\b(bl\s*date|movement\s*date|suppliers?\s*invoice\s*date)\b", 10, "amend BL/movement/supplier date"),
        (r"\b(incorrect|wrong|different)\b.{0,30}\b(invoice\s*date|posting\s*date|date)\b", 10, "incorrect date"),
        (r"\bperiod\s*1\b", 10, "period 1"),
        (r"\bchange\s*(processing|account\s*posting)\s*date\b", 10, "change processing date"),
        (r"\b(cannot|unable)\s*to\s*change\s*bl\s*date\b", 10, "cannot change BL date"),
        (r"\b(change|amend|update)\b.{0,20}\bdue\s*date\b", 9, "change due date"),
        (r"\btransfer\s*doc\b.{0,30}\b(from\s*)?(ne\s*)?(month|november|december|january|february|march)\b", 9, "transfer doc to NE month"),
    ],

    "Discrepancies - Tax": [
        (r"\b" + VERBS_AMEND + r"\b.{0,30}\b(tax|vat|gst|withholding|sst)\b", 11, "amend/update tax"),
        (r"\b(tax|vat|gst|withholding|sst)\b.{0,30}\b(description|code|event|type|ruling)\b", 10, "tax description/code/event"),
        (r"\badd(ed)?\s*tax\s*line\b|\badd.{0,10}tax\s*fee\b", 10, "add tax line/fee"),
        (r"\b(tax\s*(type|code|line|ruling))\s*(not\s*correct|incorrect|wrong|check|type\s*modification\s*fail)\b", 10, "tax code/type incorrect"),
        (r"\bcurrent\s*tax\s*type\s*(has\s*changed|doesn.?t\s*match|changed)\b", 11, "tax type changed/mismatch"),
        (r"\btax\s*(line|code)\s*(did\s*not\s*appear|not\s*appear|missing)\b", 11, "tax line/code not appearing"),
        (r"\bno\s*tax\s*line\b", 12, "no tax line"),
        (r"\btax\s*line\s*(not\s*appear|missing|did\s*not\s*appear|does\s*not\s*appear)\b", 12, "tax line not appearing"),
        (r"\bmissing\s*tax\s*line\s*item\b", 12, "missing tax line item"),
        (r"\btax\s*code\s*did\s*not\s*reflect\b", 12, "tax code did not reflect"),
        (r"\b(no|without)\s*tax\s*type\b", 11, "no tax type"),
        (r"\b(tax\s*overwrite|override\s*tax)\b", 10, "tax overwrite"),
        (r"\bstuck\s*at\s*(sent\s*for\s*tax\s*overwrite|trigger\s*tax)\b", 11, "stuck at tax overwrite"),
        (r"\bcannot\s*process\b.{0,30}\b(tax|override)\b", 10, "cannot process due to tax"),
        (r"\b2\s*taxline\b|\btwo\s*tax\s*line\b", 10, "2 tax lines appeared"),
    ],

    "Discrepancies - Doc": [
        (r"\b" + DOC_TERMS + r"\b.{0,20}\b(number|no)\b.{0,20}\b" + VERBS_AMEND + r"\b", 10, "amend invoice/document number"),
        (r"\b(amend|fix|correct)\b.{0,20}\b(doc(ument)?|invoice)\b", 9, "amend doc/invoice"),
        (r"\b(reject|cancel)\b.{0,20}\b(invoice|ifp|doc(ument)?)\b", 8, "reject/cancel invoice/IFP"),
        (r"\b(wrong|incorrect)\s*(doc\s*type|document\s*type|sap\s*doc\s*type)\b", 10, "wrong doc type"),
        (r"\b(invoice\s*number|doc\s*number|running\s*number)\b.{0,20}\b(wrong|incorrect|amend)\b", 10, "wrong invoice/doc number"),
        (r"\b(cancel|reverse)\b.{0,20}\b(invoice|ifp|cn|dn)\b", 8, "cancel/reverse invoice"),
        (r"\bstrategy\b.{0,20}\bnot\s*combined\b", 9, "strategy not combined"),
        (r"\bcn\s*doc\s*ref\b.{0,30}\bwrong\b", 9, "CN doc ref wrong"),
        (r"\bincorrect\s*(credit|debit)\s*(or\s*debit\s*)?(entry|entries|note)\b", 9, "incorrect credit/debit entry"),
        (r"\baccounting\s*type\b.{0,20}\b(wrong|incorrect|change|amend)\b", 9, "accounting type wrong"),
        (r"\bpayment\s*method\b.{0,30}\b(wrong|incorrect|reflected\s*as|amend|change)\b", 9, "payment method wrong"),
        (r"\b(issue|process|to\s*issue)\s*(100%\s*)?(credit\s*note|cn)\b", 8, "issue credit note"),
        (r"\brevert\s*(the\s*)?status\b.{0,30}\b(back\s*to|to)\b.{0,30}\b(reserve|draft|planned)\b", 8, "revert status"),
        (r"\baccidentally\s*(changed|deleted|removed)\b.{0,30}\b(status|deal|doc)\b", 8, "accidentally changed status/deal"),
        (r"\b(physical|cash)\s*settlement\b.{0,20}\b(wrong(ly\s*tagged)?|instead\s*of|should\s*be)\b", 9, "settlement type wrong"),
        (r"\bchange\s*p\d\s*to\s*p\d\b", 8, "change provisional level"),
    ],

    "Discrepancies - Account": [
        (r"\b(gl|account|coa)\b.{0,20}\b" + VERBS_AMEND + r"\b", 9, "amend GL/account"),
        (r"\b(correct|wrong|incorrect)\b.{0,20}\b(gl|account|coa)\b", 9, "wrong GL/account"),
        (r"\b(gl|account)\s*(mapping|should\s*be|is\s*wrong|incorrect)\b", 9, "GL mapping wrong"),
        (r"\bsap\s*doc\s*type\b.{0,30}\b(not\s*auto.?pop|missing|not\s*appear|wrong)\b", 10, "SAP doc type missing/wrong"),
        (r"\btag\s*gl\s*account\b.{0,20}\b(fee|type)\b", 9, "tag GL account fee type"),
        (r"\baccount\s*name\b.{0,20}\b(should\s*be|change\s*to|wrong|incorrect|i/o)\b", 8, "account name wrong"),
        (r"\bmanually\s*change\b.{0,20}\b(accounting\s*type|from\s*v\s*to\s*g|from\s*g\s*to\s*v)\b", 9, "manually change accounting type"),
    ],

    "Discrepancies - Bank Account Issue": [
        (BANK_TERMS, 11, "bank account/IBAN/SWIFT"),
        (r"\b(add|update|amend)\s*(account\s*number|acct\s*no|myr\s*account)\b", 10, "add/update account number"),
        (r"\b(beneficiary|remittance\s*bank|bank\s*address)\b.{0,30}\b(missing|not\s*appear|incorrect|no\s*dropdown)\b", 10, "beneficiary/bank details missing"),
        (r"\b(external|internal)\s*settle\s*instruction\b.{0,30}\b(not\s*appear|wrong|missing|none)\b", 10, "settle instruction missing/wrong"),
        (r"\bsettle\s*instruction\b.{0,30}\b(not\s*appear|none|incorrect)\b", 10, "settle instruction none"),
        (r"\b(add|include|insert)\s*bank\s*(details?|acc(ount)?)\b.{0,30}\b(counterpart|endur|imos)\b", 9, "add bank details counterpart"),
        (r"\b(add|include|insert)\s*va\b.{0,30}\b(counterpart|endur|imos)\b", 9, "add VA counterpart"),
        (r"\bva\b.{0,20}\b(not\s*(listed|in|found)|add|register)\b.{0,20}\b(imos|endur|dropdown)\b", 9, "VA not in IMOS/Endur"),
        (r"\bremove\b.{0,20}\bbank\s*address\b", 9, "remove bank address"),
        (r"\b(add|register)\s*(receivable|payable)\s*va\b", 9, "add receivable/payable VA"),
    ],

    "Discrepancies - Customer Code": [
        (r"\bcustomer\s*code\b.{0,20}\b" + VERBS_AMEND + r"\b", 10, "amend customer code"),
        (r"\bcustomer\s*(code|number)\s*(wrong|incorrect|not\s*correct|not\s*reflect)\b", 9, "wrong customer code"),
        (r"\breset\s*customer\s*code\b", 10, "reset customer code"),
        (r"\binsert\s*(the\s*)?correct\s*customer\s*code\b", 10, "insert correct customer code"),
    ],
    "Discrepancies - Company Code": [
        (r"\bcompany\s*code\b.{0,20}\b" + VERBS_AMEND + r"\b", 10, "amend company code"),
        (r"\b(manually\s*)?change\s*(the\s*)?sap\s*company\s*code\b", 10, "change SAP company code"),
    ],
    "Discrepancies - Legal Entity": [
        (r"\blegal\s*entity\b.{0,20}\b" + VERBS_AMEND + r"\b", 10, "amend legal entity"),
        (r"\b(wrong|incorrect)\b.{0,20}\blegal\s*entity\b", 9, "wrong legal entity"),
    ],
    "Discrepancies - Broker broker": [
        (r"\b(broker|fee\s*broker)\b.{0,20}\b" + VERBS_AMEND + r"\b", 9, "amend broker/fee broker"),
        (r"\bdelete\b.{0,20}\b-\s*broker\b", 9, "delete -BROKER suffix"),
    ],
    "Discrepancies - Strategy": [
        (r"\bstrategy\b.{0,20}\b" + VERBS_AMEND + r"\b", 9, "amend strategy"),
        (r"\b(wrong|incorrect|not\s*correct)\b.{0,20}\bstrategy\b", 8, "wrong strategy"),
        (r"\bretag\s*(to\s*)?(a\s*)?different\s*strategy\b", 9, "retag strategy"),
        (r"\bactivity\s*name\b.{0,20}\b(incorrect|wrong|should\s*be)\b", 8, "activity name incorrect"),
    ],
    "Discrepancies - Address": [
        (r"\baddress\b.{0,20}\b" + VERBS_AMEND + r"\b", 9, "amend address"),
        (r"\bcorrect\s*(the\s*)?spelling\b.{0,20}\b(endur|trms|system|counterpart)\b", 7, "correct spelling"),
    ],
    "Discrepancies - vessel": [
        (r"\bvessel\b.{0,20}\b" + VERBS_AMEND + r"\b", 9, "amend vessel"),
        (r"\bvessel\s*name\b.{0,30}\b(wrong|incorrect|not\s*appear|missing|change)\b", 9, "vessel name wrong/missing"),
        (r"\badd\s*vessel\s*name\b|\bchange\b.{0,20}\bvessel\s*name\b", 9, "add/change vessel name"),
    ],
    "Discrepancies - Freight": [
        (r"\bfreight\b.{0,20}\b" + VERBS_AMEND + r"\b", 9, "amend freight"),
        (r"\bfreight\s*(amount|cost|fee)\b.{0,30}\b(not\s*appear|missing|wrong|incorrect)\b", 9, "freight amount missing/wrong"),
        (r"\bfreight\s*segregation\b|\bfreight\b.{0,30}\bwrongly\s*tagged\b", 9, "freight segregation/wrongly tagged"),
        (r"\bincorrect\s*freight\s*fee\b|\bfreight\b.{0,30}\b(50[/\\]50|100%|percentage|share)\b", 8, "freight fee issue"),
    ],
    "Discrepancies - Bunker": [
        (r"\bbunker\b.{0,20}\b" + VERBS_AMEND + r"\b", 9, "amend bunker"),
        (r"\bbunker\s*(expenses?|cost|fee|amount)\b.{0,30}\b(not\s*available|not\s*appear|missing)\b", 9, "bunker expenses missing"),
        (r"\bbunker\s*(price|adjustment)\b.{0,20}\b(not|wrong|incorrect)\b", 9, "bunker price/adjustment issue"),
    ],
    "Discrepancies - Counterparty": [
        (r"\b(counterparty|vendor|supplier)\b.{0,20}\b" + VERBS_AMEND + r"\b", 9, "amend counterparty/vendor"),
        (r"\b(vendor\s*code)\s*(not\s*in\s*endur|register|backend|amend|update)\b", 9, "vendor code update"),
        (r"\bchange\s*(the\s*)?name\s*(from|of)\b.{0,40}\b(to|pte|ltd|llc|corp)\b", 8, "change counterparty name"),
        (r"\bchange\s*(vendor\s*code|the\s*vendor\s*code)\b|\badd\s*vendor\s*code\b", 9, "change/add vendor code"),
        (r"\b(counterparty|counterpart|vendor|supplier)\b.{0,20}\bnot\s*(in|listed|found)\b.{0,20}\b(endur|system|imos)\b", 9, "counterparty not in Endur"),
    ],
    "Discrepancies - Details": [
        (r"\bdetails?\b.{0,20}\b" + VERBS_AMEND + r"\b", 8, "amend details"),
        (r"\b(issuer\s*name|issuer)\b.{0,20}\b(change|amend|wrong|different)\b", 8, "change issuer name"),
    ],

    "Question Help Support": [
        (r"\bhow\s*to\b|\bkindly\s*advise\b|\bplease\s*advise\b|\bneed\s*guidance\b", 8, "asks how to/advise"),
        (r"\b(please|kindly)\s*(assist\s*to\s*(check|confirm|verify|explain|clarify|look|investigate|provide))\b", 7, "assist to check/verify"),
        (r"\bwould\s*like\s*to\s*(understand|explore|know|request)\b", 7, "seeking understanding"),
        (r"\bcan\s*(you|trms)\s*(advise|confirm|explain|check|clarify)\b", 7, "can you advise/confirm"),
        (r"\b(how\s*do\s*i|what\s*should\s*i|what\s*are\s*the\s*steps?|next\s*steps?)\b", 7, "asking steps/process"),
        (r"\b(change|update|revert)\s*(loa|doa)\b", 9, "LOA/DOA change request"),
        (r"\b(loa|doa)\b.{0,30}\b(change|from|to|leave|effective)\b", 8, "LOA/DOA from/to"),
        (r"\b(add|request\s*to\s*add|include)\s*(verifier|approver)\b.{0,20}\b(endur|imos|trms)\b", 8, "add verifier/approver"),
        (r"\brequest\s*to\s*add\s*(verifier|approver|designation|department)\b", 8, "request to add verifier"),
        (r"\b(new\s*hire|new\s*joiner|new\s*staff)\b.{0,30}\b(endur|imos|trms|setup|department)\b", 8, "new hire setup"),
        (r"\b(create|kindly\s*create|please\s*create)\s*(a\s*new\s*)?(sn|strategy)\b", 8, "create strategy"),
        (r"\bnew\s*(strategy|sn)\b.{0,20}\b(create|business\s*unit|bu)\b", 8, "new strategy create"),
        (r"\b(add|create|make)\s*(ticker|contract\s*month|murban)\b", 8, "add/create ticker"),
        (r"\b(add|include)\s*cost\s*cent(re|er)s?\b.{0,20}\b(imos|endur|ptuk)\b", 7, "add cost centres IMOS"),
        (r"\bextend\b.{0,20}\b(vendor|counterpart|supplier)\b.{0,20}\b(imos|endur)\b", 7, "extend vendor IMOS"),
        (r"\b(add|register)\s*(counterpart|counterparty|va)\b.{0,30}\b(imos|endur)\b", 8, "add counterpart IMOS"),
        (r"\bkill\s*(my\s*)?(session|endur)\b", 7, "kill session"),
        (r"\botc\s*functionality\b|\b(add|include)\s*blending\s*component\b", 7, "OTC/blending"),
        (r"\b(add|please\s*add)\s*conversion\s*factor\b", 7, "add conversion factor"),
        (r"\b(add|change|revert|nominate)\b.{0,20}\b(approver|aa)\b.{0,20}\b(list|endur|imos|back)\b", 7, "add/change approver list"),
        (r"\b(unblock|off\s*script|unlock)\b.{0,20}\b(doc(ument)?|del\s*id|deal)\b.{0,20}\b(process|ifp|invoice)\b", 7, "request to unblock document"),
        (r"\bseek\s*(your\s*)?(assistance|advice)\b.{0,20}\b(way\s*forward|guidance|proceed)\b", 6, "seek advice way forward"),
    ],

    "System Error / Functional Issue": [
        (r"\bnomination\s*(save\s*)?failed\b", 12, "nomination save failed"),
        (r"\b(cargo\s*(id|status)|send\s*to\s*imos)\s*(failed|stuck|not\s*flow(ing)?)\b", 11, "cargo/IMOS failed"),
        (r"\bsend\s*to\s*imos\s*failed\b", 12, "send to IMOS failed"),
        (r"\b(error|failed)\s*(pop(s?)?\s*out|pop.?up|message|appear)\b", 9, "error pops out"),
        (r"\b(security\s*enforced|security\s*object)\b", 10, "security enforced error"),
        (r"\b(can.?t\s*(run|open|access|process)|unable\s*to\s*(run|open|access|process|save|delete|view|log\s*in|login))\b", 9, "unable to run/open/access"),
        (r"\b(endur|system|imos)\s*(not\s*responding|unresponsive|keeps?\s*running|freeze|frozen|hang(ing)?)\b", 10, "system unresponsive/hanging"),
        (r"\b(flatfile|flat\s*file)\s*(keeps?\s*running|failed\s*to\s*generate|generate\s*failed)\b", 9, "flatfile generation issue"),
        (r"\b(actuali[sz]ation|actuali[sz]e)\s*(error|failed|issue|unable)\b", 10, "actualization error"),
        (r"\b(lhdn|e.?invoice|e.?document)\s*(failed|error|not\s*work|rejected|issue)\b", 9, "LHDN/e-invoice error"),
        (r"\b(unexpected\s*error|system\s*(error|issue|bug|problem))\b", 8, "system error/bug"),
        (r"\b(unable|cannot|can.?t)\s*(process|proceed)\s*document\b", 10, "cannot process document"),
        (r"\bcargo\s*(id|status)\b.{0,30}\b(failed|not\s*flow|send\s*to\s*imos\s*failed)\b", 11, "cargo IMOS failed"),
        (r"\bpop.?up\b.{0,30}\b(error|message)\b.{0,20}\bprocess\b", 9, "popup error on process"),
        (r"\bprocess\s*(and\s*output|documents?)\b.{0,30}\b(error|fail|block|pop.?up)\b", 9, "process doc error"),
        (r"\b(cannot|unable\s*to|can.?t)\s*save\b.{0,30}\b(error|pop.?up|message)\b", 9, "cannot save error"),
        (r"\b(endur|system|imos)\s*(slow|lagging|loading|slowness|sluggish)\b", 9, "system slow/lagging"),
        (r"\b(slow|lagging|slowness)\b.{0,20}\b(endur|system|imos|process|load)\b", 9, "slow system"),
        (r"\btaking\s*(too\s*long|forever|long\s*time)\b.{0,20}\b(endur|save|load|process|respond)\b", 9, "taking too long"),
        (r"\bunable\s*to\s*actuali[sz]e\b|\bcannot\s*actuali[sz]e\b", 10, "unable to actualize"),
        (r"\b(qr\s*code|lhdn\s*qr)\s*(not\s*(appear|generat|receiv)|failed|missing)\b", 9, "QR code not generated"),
        (r"\bdoc(ument)?\s*(data\s*)?generation\s*fail(ure|ed)?\b", 9, "document generation failure"),
        (r"\b(endur|system|trms|imos)\s*(crash(ed)?|crashing)\b|\bcrash\b.{0,20}\b(endur|system|trms|imos)\b", 10, "system crashed"),
        (r"\bconnection\s*interrupted\b|\b(cannot|unable)\s*to\s*connect\b", 8, "connection interrupted"),
        (r"\b(cannot|unable|not\s*able)\s*to\s*change\s*(the\s*)?status\b", 9, "cannot change status"),
        (r"\b(couldn.?t|cannot|can.?t|unable\s*to)\s*save\b.{0,30}\b(cargo|nomination|shipment|changes)\b", 9, "cannot save cargo/nomination"),
        (r"\b(issue|problem|cannot|unable)\b.{0,20}\b(confirm\s*plan|planned\s*to\s*actual)\b", 9, "cannot confirm plan"),
        (r"\b(cmotion|c\s*motion)\s*workbench\b.{0,30}\b(not\s*work|issue|error|no\s*loading)\b", 9, "cMotion workbench issue"),
        (r"\b(approver|aa)\b.{0,20}\b(not\s*(appear|work|valid)|error|invalid|no\s*longer)\b", 9, "approver issue"),
        (r"\b(not\s*receive|didn.?t\s*receive|no)\s*(notification|email)\b.{0,30}\b(approval|approved|aa)\b", 8, "no notification after approval"),
        (r"\b(keeps?\s*(loading|running)|infinite\s*load|stuck\s*(on|at)\s*(loading|drafted|status))\b", 9, "infinite loading"),
        (r"\b(can.?t|cannot|unable)\s*(open|start|fire\s*up)\s*(endur|imos)\b|\b(endur|imos)\b.{0,20}\b(can.?t|cannot|unable)\s*to\s*(open|start)\b", 9, "cannot open endur"),
        (r"\bavailable\s*disk\s*space\b|\b(disk\s*(m|data|space)|drive\s*m)\b.{0,20}\b(not\s*available|missing|no\s*longer)\b", 7, "disk space issue"),
        (r"\bimos\b.{0,20}\b(very\s*slow|slow\s*loading|running\s*slowly)\b", 9, "IMOS very slow"),
        (r"\b(document|doc)\b.{0,20}\b(stuck\s*at|stuck\s*in)\b.{0,20}\b(tpm|approval|approve|status)\b", 9, "doc stuck at TPM/approval"),
        (r"\b(saving|save)\s*(process|takes)\b.{0,30}\b(delay|15|20\s*minutes?|slow|long)\b", 9, "saving process slow"),
        (r"\b(zero(rise|rize)|zerorize)\b.{0,30}\b(cannot|unable|issue|error)\b|\b(cannot|unable)\b.{0,20}\b(zero(rise|rize)|zerorize)\b", 8, "cannot zerorize"),
        (r"\b(endur|system)\b.{0,30}\b(saving|save)\b.{0,30}\b(15|20|30)\s*minutes?\b", 9, "endur saving too long"),
        (r"\bparcel\b.{0,20}\b(inconsistent|incorrect|issue|problem|error)\b", 7, "parcel data inconsistent"),
        (r"\bposition\s*conversion\b.{0,20}\b(off|incorrect|wrong)\b|\b(apm|risk\s*engine)\b.{0,20}\b(off|incorrect|wrong|issue)\b", 7, "position conversion/APM issue"),
    ],

    "Report Issue": [
        (r"\b(eod|end\s*of\s*day)\s*(report|pnl|p&l|grm)\b.{0,40}\b(wrong|incorrect|missing|not\s*appear|not\s*include)\b", 11, "EOD report wrong/missing"),
        (r"\b(pnl|p&l|pnl_cy|pnl_ly|var|value\s*at\s*risk)\b.{0,20}\b(report|breakdown)\b.{0,40}\b(wrong|incorrect|missing|issue|discrepancy)\b", 10, "PnL/VaR report issue"),
        (r"\b(grm|pnl\s*breakdown|raw\s*data\s*detail|report\s*builder)\b.{0,40}\b(wrong|incorrect|missing|issue|not\s*work|cannot\s*run)\b", 10, "GRM/PnL Breakdown report issue"),
        (r"\b(cannot|unable)\s*(to\s*)?(run|generate|produce)\b.{0,20}\b(report|grm|pnl)\b", 10, "unable to run report"),
        (r"\brealised\s*derivatives\s*report\b.{0,40}\b(wrong|incorrect|missing|not\s*appear)\b", 11, "Realised Derivatives report issue"),
        (r"\b(var|value\s*at\s*risk)\s*(report|decomposed|intra.?day)\b.{0,40}\b(wrong|incorrect|missing)\b", 10, "VaR report issue"),
        (r"\b(pnl|p&l)\s*(year\s*)?bucket\b.{0,30}\b(wrong|incorrect|not\s*correct)\b|\bwrong\s*(year\s*bucket|pnl\s*bucket)\b", 10, "PnL year bucket wrong"),
        (r"\bderivatives\s*p&l\s*report\b", 11, "derivatives P&L report"),
        (r"\bcategory\s*of\s*issue\s*[:]\s*wrong\s*(p&l|pnl|price|quantity)\b", 12, "derivatives report category"),
        (r"\bcannot\s*(load|extract|generate|download)\b.{0,20}\breport\b", 9, "cannot load report"),
        (r"\bflatfile\b.{0,30}\b(appeared\s*in\s*incorrect\s*tab|wrong\s*tab|incorrect\s*entitlement)\b", 9, "flatfile wrong tab"),
        (r"\bpnl_cy\b.{0,20}\b(zero|0|column|missing|no)\b|\bno\s*pnl_cy\b", 9, "pnl_cy zero"),
        (r"\b(balance\s*deal|new\s*balance\s*deal)\b.{0,30}\b(incorrect|wrong|seem|weird|unexpected)\b", 8, "balance deals incorrect"),
        (r"\bstrateg\w*\b.{0,20}\bshowing\s*new\s*balance\s*deals?\b", 8, "strategies showing new balance deals"),
    ],

    "Access Issue": [
        (r"\b(unable|can.?t)\s*to?\s*(log\s*(in|into)|login|access|open)\b.{0,30}\b(endur|imos|system|citrix|desktop|trms)\b", 11, "unable to login/access system"),
        (r"\b(citrix|imos\s*id)\b.{0,20}\b(access|removed|not\s*available|issue|problem)\b", 10, "Citrix/IMOS access issue"),
        (r"\b(access\s*issue|login\s*issue|cannot\s*login)\b", 10, "access/login issue"),
        (r"\b(imos\s*id)\s*(had\s*been|was|is)\s*(removed|deleted|revoked)\b", 11, "IMOS ID removed"),
        (r"\b(failed\s*to\s*login|unable\s*to\s*log\s*in|cannot\s*log\s*in|login\s*fail(ure|ed)?)\b", 11, "login failure"),
        (r"\bfailed\s*to\s*connect\b.{0,30}\b(server|citrix|endur)\b", 10, "failed to connect server"),
        (r"\b(unable|cannot|can.?t)\s*(access|open|launch|load)\s*(endur|citrix|imos)\b", 11, "cannot access endur/citrix"),
        (r"\b(imos\s*id|endur\s*id)\b.{0,30}\b(long|medical|maternity)?\s*leave\b", 9, "IMOS ID leave"),
        (r"\breactivate\b.{0,20}\b(imos|endur)\s*(id|account)\b", 9, "reactivate IMOS/ENDUR ID"),
        (r"\b(trouble|problem|issue)\b.{0,20}\b(log(ging)?\s*(in|into)|login)\b.{0,20}\b(endur|trms|imos)\b", 10, "trouble logging in"),
        (r"\bhaving\s*issues?\s*(logging|log\s*in|login)\b", 10, "having issues logging in"),
        (r"\b(keeps?\s*denied|auto\s*close\s*the\s*application)\b", 10, "keeps denied/auto close"),
        (r"\b(no\s*permission|do\s*not\s*have\s*permission)\b.{0,30}\b(view|download|access|see)\b", 9, "no permission to view"),
        (r"\b(i\s*)?(can.?t|cannot|unable)\s*(open|start|access)\s*(my\s*)?(endur|imos|trms)\b", 10, "I cannot open endur"),
        (r"\b(i\s*)?(have|am\s*having)\s*(trouble|problem|issue)\s*(logging|log\s*in|with\s*endur)\b", 9, "have trouble logging"),
    ],

    "PLSB Bucket Amendment": [
        (r"\bplsb\s*(year\s*)?bucket\b", 12, "PLSB bucket keyword"),
        (r"\b(year\s*bucket|bucket(ing)?)\s*(2025|2026|from\s*20\d{2}\s*to\s*20\d{2})\b", 11, "year bucket 2025/2026"),
        (r"\b(amend|change|update|transfer)\b.{0,30}\b(plsb|year\s*bucket|bucket(ing)?)\b", 11, "amend PLSB/year bucket"),
        (r"\bnot\s*applicable\b.{0,20}\b(bucket|plsb)\b|\b(bucket|plsb)\b.{0,20}\bnot\s*applicable\b", 10, "Not Applicable bucket"),
        (r"\b(transfer|move|change)\b.{0,20}\bbucket\b.{0,20}\b20(25|26)\b", 10, "transfer bucket year"),
        (r"\bfee\s*pv\b.{0,20}\b(not\s*auto.?calculat|does\s*not\s*auto|issue)\b", 9, "fee PV not auto calculated"),
    ],

    "Matching Issue": [
        (r"\b(unable\s*to\s*match|cannot\s*match|fail(ed)?\s*to\s*match)\b", 12, "unable to match"),
        (r"\b(matching\s*(issue|problem|error|failed))\b", 11, "matching issue"),
        (r"\b(deal|cargo)\s*(not\s*(matched|matche[sd])|unmatched)\b", 10, "deal not matched"),
        (r"\b(break\s*match|breakmatch)\b", 10, "break match"),
        (r"\b(unable\s*to|cannot|can.?t)\s*(break\s*match|unmatche)\b", 10, "unable to break match"),
        (r"\b(actuali[sz](e|ation))\b.{0,30}\b(error|fail(ed)?|unable|issue)\b", 11, "actualize error variant"),
        (r"\berror\b.{0,30}\b(during|when|while)\b.{0,20}\bmatching\b|\bmatching\b.{0,20}\berror\b", 10, "error during matching"),
        (r"\b(nomination\s*(save\s*)?failed)\b.{0,30}\bmatch", 11, "nomination save failed during match"),
        (r"\binsufficient\b.{0,20}\b(balance|storage|tank)\b", 10, "insufficient balance"),
        (r"\bvolume\s*calc\s*type\b.{0,30}\bnomination\s*save\s*failed\b", 11, "volume calc type nomination failed"),
        (r"\b(issue\s*with\s*rematching|re.?match\s*(deal|issue))\b", 9, "rematch issue"),
        (r"\bbreak\s*match\w*\b.{0,30}\b(voyage|cargo\s*id|imos)\b", 9, "break match with voyage"),
    ],

    "New Setup / Configuration": [
        # Registering new entities in the system
        (r"\b(register|onboard|add|create|setup|set\s*up)\b.{0,30}\b(new\s*)?(counterpart(y|ies)?|vendor|supplier)\b.{0,30}\b(endur|imos|system|backend|trms)\b", 11, "register new counterparty"),
        (r"\b(add|create|setup|set\s*up)\b.{0,25}\b(new\s*)?(ticker|instrument|contract\s*month|product)\b", 10, "add new ticker/instrument"),
        (r"\b(create|add|setup)\b.{0,20}\b(new\s*)?(strategy|sn|business\s*unit|bu)\b.{0,30}\b(endur|imos|trms|system)\b", 10, "create new strategy"),
        (r"\b(add|extend|register)\b.{0,20}\b(vendor|counterpart)\b.{0,20}\b(to\s*)?(imos|endur)\b", 10, "extend vendor to IMOS"),
        (r"\b(add|register)\b.{0,20}\b(new\s*(hire|staff|joiner|user))\b", 9, "new hire onboarding"),
        (r"\b(setup|set\s*up|configure|provision)\b.{0,30}\b(endur|imos|trms)\b.{0,30}\b(access|account|id|user)\b", 9, "provision system access"),
        (r"\b(new\s*)?(loa|doa)\b.{0,20}\b(setup|create|add|effective|from)\b", 9, "LOA/DOA setup"),
        (r"\b(add|create|please\s*add)\s*(conversion\s*factor|new\s*sn|new\s*product|blending\s*component)\b", 9, "add conversion/sn/product"),
        (r"\b(add|include)\b.{0,20}\b(cost\s*cent(re|er)|profit\s*cent(re|er))\b.{0,20}\b(imos|endur)\b", 9, "add cost/profit centre"),
        (r"\bnew\s*(ticker|contract\s*month|delivery\s*batch|product)\b.{0,30}\b(to\s*)?(system|endur|imos|list)\b", 10, "new config item in system"),
        (r"\b(configure|configuration)\b.{0,25}\b(not\s*found|missing|required|setup)\b", 8, "configuration missing"),
    ],
}

# ============================================================
# EXCEL-DRIVEN KEYWORD MANAGEMENT
# ============================================================
# Non-technical users edit Classifier_Keywords.xlsx — no code needed.
#
# Pattern column uses plain readable phrases:
#
#   Single phrase     →  wrong price
#   Either/or (|)     →  wrong price | incorrect price
#   Both words (+)    →  wrong + price        (both must appear near each other)
#   Exact phrase      →  "strike price"       (quote for exact multi-word match)
#
# Score guide:  12-15 = very strong match
#               9-11  = strong match
#               6-8   = medium hint
#               1-5   = weak hint
#
# Enabled: YES to use the rule, NO to disable without deleting.
# ============================================================

KEYWORDS_FILE = OUT_DIR / "Classifier_Keywords.xlsx"


def _regex_from_readable(raw: str) -> str:
    """
    Convert a plain-English pattern phrase into a regex string.

    Supported formats (case-insensitive, all converted to regex):
      wrong price              → \bwrong\s+price\b
      wrong price | bad price  → (\bwrong\s+price\b|\bbad\s+price\b)
      wrong + price            → \bwrong\b.{0,60}\bprice\b  (proximity AND)
      "strike price"           → \bstrike\s+price\b  (exact phrase)
    """
    raw = raw.strip()

    # Already a regex — contains regex special syntax, pass through
    if any(c in raw for c in [r"\b", ".{", "(?", "(?i", "(?s"]):
        return raw

    # OR operator — split on |
    if "|" in raw:
        parts = [p.strip() for p in raw.split("|") if p.strip()]
        return "(" + "|".join(_regex_from_readable(p) for p in parts) + ")"

    # AND proximity operator — split on +
    if "+" in raw:
        parts = [p.strip() for p in raw.split("+") if p.strip()]
        # Each part becomes a word-boundary match, joined with up to 60 chars between
        compiled_parts = []
        for p in parts:
            words = p.strip().strip('"').split()
            compiled_parts.append(r"\b" + r"\s+".join(re.escape(w) for w in words) + r"\b")
        return r".{0,60}".join(compiled_parts)

    # Exact phrase in quotes
    if raw.startswith('"') and raw.endswith('"'):
        phrase = raw[1:-1].strip()
        words  = phrase.split()
        return r"\b" + r"\s+".join(re.escape(w) for w in words) + r"\b"

    # Plain phrase — wrap whole phrase in word boundaries
    words = raw.split()
    if len(words) == 1:
        return r"\b" + re.escape(words[0]) + r"\b"
    return r"\b" + r"\s+".join(re.escape(w) for w in words) + r"\b"


def _readable_from_regex(pat: str, desc: str) -> str:
    """
    Convert a regex pattern to a readable phrase for the Excel export.
    Uses the description as the readable label — much cleaner than raw regex.
    Falls back to a simplified version of the pattern if no description.
    """
    if desc and desc.strip():
        return desc.strip()
    # Strip common regex noise for a rough readable version
    readable = pat
    readable = re.sub(r"\\b", "", readable)
    readable = re.sub(r"\\.{0,\d+}\\b", " ... ", readable)
    readable = re.sub(r"\(\?[is]+\)", "", readable)
    readable = re.sub(r"\\s\*", " ", readable)
    readable = re.sub(r"\\s\+", " ", readable)
    readable = re.sub(r"\(([^|()]{1,30})\)", r"\1", readable)
    readable = re.sub(r"\s+", " ", readable).strip()
    return readable[:80]


def export_keywords_to_excel(path: Path, patterns: dict, cats: list) -> None:
    """
    Write all patterns to Excel using plain readable phrases.
    The Description column becomes the Pattern column — human-readable.
    """
    rows = []
    for cat in cats:
        if cat == "Others":
            continue
        for pat_str, score, desc in patterns.get(cat, []):
            readable = _readable_from_regex(pat_str, desc)
            rows.append({
                "Category":    cat,
                "Pattern":     readable,
                "Score":       score,
                "Description": desc,
                "Enabled":     "YES",
            })

    df_kw = pd.DataFrame(rows, columns=["Category", "Pattern", "Score", "Description", "Enabled"])

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df_kw.to_excel(writer, sheet_name="Keywords", index=False)

        # Format the Keywords sheet
        from openpyxl.styles import PatternFill, Font, Alignment
        wb  = writer.book
        ws  = writer.sheets["Keywords"]
        hdr_fill = PatternFill("solid", fgColor="0D1B2A")
        hdr_font = Font(color="FFFFFF", bold=True, name="Calibri", size=10)
        for cell in ws[1]:
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions["A"].width = 35   # Category
        ws.column_dimensions["B"].width = 45   # Pattern
        ws.column_dimensions["C"].width = 8    # Score
        ws.column_dimensions["D"].width = 40   # Description
        ws.column_dimensions["E"].width = 10   # Enabled
        ws.freeze_panes = "A2"

        # Instructions sheet
        instructions = pd.DataFrame({
            "Column": [
                "Category", "Pattern", "Score", "Description", "Enabled",
                "", "HOW TO WRITE PATTERNS", "",
                "Example", "What it does",
            ],
            "Explanation": [
                "The issue category. Must match exactly (e.g. Wrong Price).",
                "Plain words to search for in ticket descriptions. See examples below.",
                "Points added when this rule matches. 12-15 = strong, 9-11 = medium, 6-8 = weak.",
                "Your own label for this rule. Does not affect matching — just for reference.",
                "YES = active. NO = disabled (rule is ignored but not deleted).",
                "", "", "",
                "wrong price",               "Matches tickets containing the words 'wrong price'",
            ],
        })
        instructions.to_excel(writer, sheet_name="Instructions", index=False)

        ws2 = writer.sheets["Instructions"]
        # Add more example rows manually
        examples = [
            ("wrong price | incorrect price",  "Matches 'wrong price' OR 'incorrect price'"),
            ("wrong + price",                  "Matches tickets where 'wrong' AND 'price' both appear (within 60 chars)"),
            ('"strike price"',                 "Matches the exact phrase 'strike price'"),
            ("push + NE",                      "Matches tickets where 'push' AND 'NE' both appear near each other"),
            ("",                               ""),
            ("TIPS", ""),
            ("Use | to match multiple phrases for the same rule",  ""),
            ("Use + when two words must both appear but not necessarily next to each other", ""),
            ("Higher score = wins when two categories tie",        ""),
            ("Set Enabled = NO to test removing a rule safely",    ""),
        ]
        start_row = len(instructions) + 3
        for i, (ex, expl) in enumerate(examples):
            ws2.cell(row=start_row + i, column=1, value=ex)
            ws2.cell(row=start_row + i, column=2, value=expl)

        ws2.column_dimensions["A"].width = 50
        ws2.column_dimensions["B"].width = 60

    print(f"  ✅ Keywords file created → {path}")
    print(f"     Open Classifier_Keywords.xlsx to add/edit/disable keyword rules.")
    print(f"     No code changes needed — just edit and save the Excel file.")


def load_patterns_from_excel(path: Path) -> "tuple[dict, list] | None":
    """
    Load PATTERNS and CATS from the keywords Excel file.
    Pattern column is plain readable text — converted to regex internally.
    """
    try:
        df_kw = pd.read_excel(path, sheet_name="Keywords", engine="openpyxl")
    except Exception as e:
        print(f"  [WARN] Could not read keywords file: {e}")
        return None

    required = {"Category", "Pattern", "Score"}
    if not required.issubset(set(df_kw.columns)):
        print(f"  [WARN] Keywords file missing required columns {required} — using built-in patterns.")
        return None

    if "Enabled" in df_kw.columns:
        df_kw = df_kw[df_kw["Enabled"].astype(str).str.upper().str.strip() != "NO"].copy()

    df_kw = df_kw.dropna(subset=["Category", "Pattern", "Score"])
    df_kw["Score"]       = pd.to_numeric(df_kw["Score"], errors="coerce").fillna(8).astype(int)
    df_kw["Category"]    = df_kw["Category"].astype(str).str.strip()
    df_kw["Pattern"]     = df_kw["Pattern"].astype(str).str.strip()
    df_kw["Description"] = df_kw["Description"].fillna("").astype(str) if "Description" in df_kw.columns else ""

    patterns_out: Dict[str, List[Tuple[str, int, str]]] = {}
    cats_out: list = []
    bad_patterns = 0

    for cat, grp in df_kw.groupby("Category", sort=False):
        if cat not in cats_out:
            cats_out.append(cat)
        rules = []
        for _, row in grp.iterrows():
            readable = str(row["Pattern"]).strip()
            desc     = str(row["Description"]).strip()
            try:
                regex = _regex_from_readable(readable)
                re.compile(regex, re.IGNORECASE)  # validate
                rules.append((regex, int(row["Score"]), desc or readable))
            except re.error as e:
                bad_patterns += 1
                print(f"  [WARN] Skipping invalid pattern in '{cat}': '{readable}' → {e}")
        patterns_out[cat] = rules

    if "Others" not in cats_out:
        cats_out.append("Others")

    if bad_patterns:
        print(f"  [WARN] {bad_patterns} pattern(s) skipped due to errors.")

    print(f"  📋 Loaded {len(df_kw):,} keyword rules from {path.name} "
          f"({len(cats_out)-1} categories)")
    return patterns_out, cats_out


# ── Load from Excel if available, otherwise use hardcoded patterns ──
if KEYWORDS_FILE.exists():
    _excel_result = load_patterns_from_excel(KEYWORDS_FILE)
    if _excel_result is not None:
        PATTERNS, CATS = _excel_result
        print("  ✅ Using Excel-driven keyword patterns.")
    else:
        print("  ⚠️  Falling back to built-in patterns.")
else:
    print(f"  📝 Keywords file not found — creating it from built-in patterns...")
    export_keywords_to_excel(KEYWORDS_FILE, PATTERNS, CATS)
    print("  ✅ Using built-in patterns for this run.")


# ── Pre-compile all patterns once at startup ──────────────────────
COMPILED: Dict[str, List[Tuple[re.Pattern, int, str]]] = {
    cat: [(re.compile(p, re.IGNORECASE), w, cue) for p, w, cue in plist]
    for cat, plist in PATTERNS.items()
}

_FB_PATTERNS = [
    # Wrong Price: require BOTH a problem word AND a price-related term
    # (removing the bare PRICE_TERMS fallback that caused over-classification).
    (re.compile(
        r"\b(wrong|incorrect|mismatch|discrepancy|not\s*tally|not\s*correct)\b.{0,60}\b"
        + PRICE_TERMS + r"\b",
        re.IGNORECASE
    ), "Wrong Price", 4),
    (re.compile(
        r"\b" + PRICE_TERMS + r"\b.{0,60}\b(wrong|incorrect|mismatch|discrepancy|not\s*tally)\b",
        re.IGNORECASE
    ), "Wrong Price", 4),
    (re.compile(AMOUNT_TERMS, re.IGNORECASE),           "Discrepancies - Amount",            3),
    (re.compile(FX_TERMS, re.IGNORECASE),               "Discrepancies - exchange rate",     3),
    (re.compile(MISSING_PHRASES, re.IGNORECASE),        "Missing Data",                      3),
    (re.compile(DATE_TERMS, re.IGNORECASE),             "Discrepancies - Date",              3),
    (re.compile(SYSTEM_ERROR_PHRASES, re.IGNORECASE),   "System Error / Functional Issue",   3),
    (re.compile(r"\b(match|matching|actuali[sz])\b", re.IGNORECASE), "Matching Issue",       3),
    (re.compile(r"\b(report|pnl|p&l|var|eod)\b", re.IGNORECASE),     "Report Issue",        3),
    (re.compile(r"\b(push|ne\s*tab|new\s*entries|je\s*tab|journal\s*entries)\b", re.IGNORECASE), "Push Document", 3),
    (re.compile(r"\bpush\s*to\s*sap\b|\bsend\s*to\s*sap\b", re.IGNORECASE), "Push to SAP", 4),
    (re.compile(r"\b(duplicate|double)\s*(line\s*item|line\s*items|barging|product\s*cost)\b", re.IGNORECASE), "run corrective", 4),
    (re.compile(r"\bi[/\\]o\b.{0,20}\b(barging|cargo|product|cost)\b", re.IGNORECASE), "run corrective", 4),
    (re.compile(DEAL_DUP_SIGNALS, re.IGNORECASE),       "Duplicate Invoice/Deal Number",     5),
    (re.compile(r"\b(reverse|reversal|reversing)\b", re.IGNORECASE),  "Discrepancies - Doc", 3),
    (re.compile(r"\b(resubmit|re.?submit)\b", re.IGNORECASE),         "SAP Posting Failures",3),
    (re.compile(r"\b(mapping|map)\b.{0,20}\b(gl|account|coa|vendor|jcc|ice)\b", re.IGNORECASE), "Discrepancies - Account", 3),
    (re.compile(r"\b(cancel|void|delete)\b.{0,20}\b(invoice|deal|document|cargo)\b", re.IGNORECASE), "Discrepancies - Doc", 3),
    (re.compile(r"\b(tax)\b", re.IGNORECASE),                          "Discrepancies - Tax",3),
    (re.compile(r"\b(invoice|inv)\b", re.IGNORECASE),                  "Discrepancies - Doc",2),
    (re.compile(r"\b(access|login|log\s*in|password|id\s*expired)\b", re.IGNORECASE), "Access Issue", 3),
    (re.compile(r"\b(corrective|fsa)\b", re.IGNORECASE),               "run corrective",     3),
    (re.compile(r"\b(loa|doa)\b.{0,30}\b(change|from|to|leave)\b", re.IGNORECASE), "Question Help Support", 3),
    (re.compile(r"\b(vendor\s*code|counterpart)\b.{0,20}\b(register|add|not\s*in|backend)\b", re.IGNORECASE), "Missing Data", 3),
    (re.compile(r"\b(slow|crash|lagging|freeze|frozen|hang)\b", re.IGNORECASE), "System Error / Functional Issue", 3),
    (re.compile(r"\b(auto.?capture|auto.?flow|autocapture|autoflow)\b", re.IGNORECASE), "Not Update Data", 3),
    (re.compile(r"\b(qr\s*(validation|code|pending)|lhdn|irb)\b", re.IGNORECASE), "Not Update Data", 3),
    (re.compile(r"\b(va\s*(number|no)?|vendor\s*account)\b.{0,20}\b(add|not\s*listed|imos|endur)\b", re.IGNORECASE), "Discrepancies - Bank Account Issue", 3),
]
_FB_MULTI_CATS = ["Wrong Price", "Discrepancies - Amount", "Wrong Quantity", "Discrepancies - Date", "Discrepancies - exchange rate"]

_INFER_PATTERNS = [
    (re.compile(PRICE_TERMS, re.IGNORECASE),     "price terms"),
    (re.compile(AMOUNT_TERMS, re.IGNORECASE),    "amount terms"),
    (re.compile(FX_TERMS, re.IGNORECASE),        "FX terms"),
    (re.compile(DATE_TERMS, re.IGNORECASE),      "date/period"),
    (re.compile(TAX_TERMS, re.IGNORECASE),       "tax terms"),
    (re.compile(MISSING_PHRASES, re.IGNORECASE), "missing/visibility"),
]

PRIORITY = [
    "run corrective", "Push Document", "Push to SAP", "SAP Posting Failures", "Duplicate Invoice/Deal Number",
    "PLSB Bucket Amendment", "Matching Issue",
    "Not Update Data", "Missing Data", "Wrong Price", "Wrong Quantity",
    "Discrepancies - Amount", "Discrepancies - exchange rate", "Discrepancies - Date",
    "Discrepancies - Tax", "Discrepancies - Doc", "Discrepancies - Account",
    "Discrepancies - Bank Account Issue", "Discrepancies - Customer Code",
    "Discrepancies - Company Code", "Discrepancies - Legal Entity",
    "Discrepancies - Broker broker", "Discrepancies - Strategy", "Discrepancies - Address",
    "Discrepancies - vessel", "Discrepancies - Freight", "Discrepancies - Bunker",
    "Discrepancies - Counterparty", "Discrepancies - Details",
    "System Error / Functional Issue", "Report Issue",
    "Question Help Support", "Access Issue", "New Setup / Configuration",
    "Others",
]
_PRIORITY_IDX = {c: i for i, c in enumerate(PRIORITY)}

def score_category(text_lc: str, cat: str) -> Tuple[int, List[str]]:
    total = 0
    cues: List[str] = []
    for rx, w, cue in COMPILED.get(cat, []):
        if rx.search(text_lc):
            total += w
            cues.append(cue)
    return total, cues

def pick_category(score_map: Dict[str, int]) -> str:
    if not score_map:
        return "Others"
    max_s = max(score_map.values())
    if max_s <= 0:
        return "Others"
    cands = [c for c, s in score_map.items() if s == max_s]
    # Check PRIORITY list first (covers hardcoded categories)
    for c in PRIORITY:
        if c in cands:
            return c
    # Excel-only categories not in PRIORITY — still works, first candidate wins
    return cands[0]

def build_reason(cat: str, cues: List[str], text_lc: str) -> str:
    cues_uniq = list(dict.fromkeys(cues))
    if not cues_uniq:
        inferred = [lbl for rx, lbl in _INFER_PATTERNS if rx.search(text_lc)][:4]
        cues_uniq = inferred
    s = f"Matched {cat} using cues: " + ", ".join(cues_uniq[:4]) + "."
    toks = s.split()
    return " ".join(toks[:40]) if len(toks) > 40 else s

def classify_text(text: str) -> tuple:
    """Return (category, reason, match_debug, confidence).

    confidence is HIGH / MEDIUM / LOW based on the winning category's score:
      HIGH   ≥ 10 points
      MEDIUM ≥  6 points
      LOW    below 6 (or empty text)
    """
    t    = norm_whitespace(text)
    # Guard against pathologically long descriptions that could slow regex
    if len(t) > 4000:
        t = t[:4000]
    t_lc = t.lower()
    if not t_lc:
        return "Others", "Low confidence / ambiguous: empty description.", "", "LOW"

    score_map: Dict[str, int]       = {}
    cues_map:  Dict[str, List[str]] = {}

    for cat in CATS:
        if cat == "Others":
            continue
        s, cues = score_category(t_lc, cat)
        if s > 0:
            score_map[cat] = s
            cues_map[cat]  = cues

    if not score_map:
        for rx, cat, pts in _FB_PATTERNS:
            if rx.search(t_lc):
                if cat == "__multi__":
                    for k in _FB_MULTI_CATS:
                        score_map[k] = score_map.get(k, 0) + pts
                else:
                    score_map[cat] = score_map.get(cat, 0) + pts

    final_cat = pick_category(score_map) if score_map else "Others"
    cues      = cues_map.get(final_cat, [])
    reason    = build_reason(final_cat, cues, t_lc)
    top3      = sorted(score_map.items(), key=lambda kv: (-kv[1], _PRIORITY_IDX.get(kv[0], 999)))[:3]
    match_dbg = "; ".join([f"{c}:{s}" for c, s in top3]) if top3 else ""

    # Confidence: based on the winning category’s raw score
    win_score  = score_map.get(final_cat, 0)
    confidence = "HIGH" if win_score >= 10 else "MEDIUM" if win_score >= 6 else "LOW"

    return final_cat, reason, match_dbg, confidence

def run_classifier() -> Path:
    print("=" * 60)
    print("PETRA — classifier.py  v4")
    print("=" * 60)
    print(f"  Input  : {INPUT_FILE}")
    print(f"  Output : {OUTPUT_FILE}")

    if not INPUT_FILE.exists():
        raise FileNotFoundError(f"Input file not found:\n  {INPUT_FILE}")

    try:
        df = pd.read_excel(INPUT_FILE, engine="openpyxl")
    except Exception as e:
        raise RuntimeError(f"Failed to read input Excel file:\n  {INPUT_FILE}\n  Error: {e}") from e

    if df.empty:
        raise ValueError(f"Input file is empty: {INPUT_FILE}")

    print(f"  Loaded {len(df):,} rows, {len(df.columns)} columns.")

    if COL_DETAILED_DESC_1 in df.columns:
        src_desc_col = COL_DETAILED_DESC_1
    elif COL_DETAILED_DESC_2 in df.columns:
        src_desc_col = COL_DETAILED_DESC_2
    else:
        src_desc_col = COL_DETAILED_DESC_1
        df[src_desc_col] = ""

    df[COL_DESC_EXTRACT] = df[src_desc_col].apply(extract_desc_only)

    for c in DATE_COLS_PREF:
        if c in df.columns:
            df[c] = to_datetime_smart(df[c])

    keys_present = [k for k in ID_COLS_PREF if k in df.columns]
    dedup_keys   = keys_present + [COL_DESC_EXTRACT]
    sort_cols    = [c for c in DATE_COLS_PREF if c in df.columns]
    if sort_cols:
        df = df.sort_values(sort_cols, ascending=[False] * len(sort_cols))

    before = len(df)
    df     = df.drop_duplicates(subset=dedup_keys, keep="first")
    print(f"  Dedup: removed {before - len(df)} rows -> {len(df):,} remain")

    print("  Classifying tickets...")
    if USE_TQDM:
        res = df[COL_DESC_EXTRACT].progress_apply(
            lambda s: pd.Series(
                classify_text(s),
                index=["Refined Summary", "Reason", "Match Debug", "Confidence"]
            )
        )
    else:
        total = len(df)
        results = []
        for i, s in enumerate(df[COL_DESC_EXTRACT], 1):
            results.append(classify_text(s))
            if i % 500 == 0 or i == total:
                print(f"    {i:,} / {total:,} rows done", flush=True)
        res = pd.DataFrame(
            results,
            columns=["Refined Summary", "Reason", "Match Debug", "Confidence"],
            index=df.index,
        )

    out_df = pd.concat([df, res], axis=1)

    if COL_PCT3 not in out_df.columns:
        out_df[COL_PCT3] = ""

    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    out_df.to_excel(OUTPUT_FILE, index=False, engine="openpyxl")

    print("\n  Coverage (Refined Summary):")
    print(out_df["Refined Summary"].value_counts(dropna=False).to_string())

    # ── Quality metrics ───────────────────────────────────────────────────
    total_rows   = len(out_df)
    others_count = (out_df["Refined Summary"] == "Others").sum()
    low_conf     = (out_df["Confidence"] == "LOW").sum()
    high_conf    = (out_df["Confidence"] == "HIGH").sum()
    print(f"\n  Classification quality:")
    print(f"    Others (unclassified) : {others_count:,} / {total_rows:,} "
          f"({others_count/total_rows*100:.1f}%)")
    print(f"    HIGH confidence       : {high_conf:,} ({high_conf/total_rows*100:.1f}%)")
    print(f"    LOW confidence        : {low_conf:,} ({low_conf/total_rows*100:.1f}%)")
    if others_count / total_rows > 0.20:
        print("    [WARN] Others > 20% — consider reviewing patterns for common ticket types.")

    print(f"\n  Classification complete -> {OUTPUT_FILE}")
    return OUTPUT_FILE


if __name__ == "__main__":
    run_classifier()
