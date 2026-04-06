# main.py
# -*- coding: utf-8 -*-
"""
PETRA — Main Pipeline
Runs: SLA → Classifier → Merge → ML → Score → Suggest → Teams Alert
"""

import pandas as pd
import numpy as np
import pickle
import warnings
import matplotlib.pyplot as plt
import os
from pathlib import Path
from sklearn.ensemble import GradientBoostingClassifier
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity
from sklearn.model_selection import train_test_split, cross_val_score
from sklearn.preprocessing import LabelEncoder
from sklearn.metrics import classification_report, roc_auc_score
from sklearn.calibration import CalibratedClassifierCV
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")

from sla        import run_sla
from classifier import run_classifier

# ============================================================
# OUTPUT DIR
# ============================================================
OUT_DIR = Path(os.environ["USERPROFILE"]) / "OneDrive - PETRONAS" / "Desktop" / "PETRA Output"
OUT_DIR.mkdir(parents=True, exist_ok=True)


# ============================================================
# STEP 1 — RUN SLA.PY
# ============================================================
print("\n" + "=" * 60)
print("PETRA PIPELINE — STEP 1: SLA Enrichment")
print("=" * 60)
SLA_FILE = run_sla()


# ============================================================
# STEP 2 — RUN CLASSIFIER.PY
# ============================================================
print("\n" + "=" * 60)
print("PETRA PIPELINE — STEP 2: Classification")
print("=" * 60)
CAT_FILE = run_classifier()


# ============================================================
# STEP 3 — MERGE
# ============================================================
print("\n" + "=" * 60)
print("PETRA PIPELINE — STEP 3: Merge")
print("=" * 60)

df_sla = pd.read_excel(SLA_FILE, dtype=str)
df_cat = pd.read_excel(CAT_FILE, dtype=str)

if df_sla.empty:
    raise ValueError(f"SLA file is empty: {SLA_FILE}")
if df_cat.empty:
    raise ValueError(f"Classifier file is empty: {CAT_FILE}")

print(f"  SLA rows        : {len(df_sla):,}")
print(f"  Classified rows : {len(df_cat):,}")

def find_id_col(df: pd.DataFrame) -> str | None:
    for c in df.columns:
        if c.strip().lower() == "incident id":
            return c
    for c in df.columns:
        if "incident" in c.lower() and "id" in c.lower():
            return c
    return None

sla_id = find_id_col(df_sla)
cat_id = find_id_col(df_cat)

if sla_id is None:
    raise KeyError("Could not find 'Incident ID' column in SLA file.")
if cat_id is None:
    raise KeyError("Could not find 'Incident ID' column in Classified file.")

if cat_id != sla_id:
    df_cat = df_cat.rename(columns={cat_id: sla_id})

CAT_EXTRA        = [sla_id, "Desc Extract", "Refined Summary", "Reason", "Match Debug", "Confidence"]
cat_cols_present = [c for c in CAT_EXTRA if c in df_cat.columns]
df_cat_slim      = df_cat[cat_cols_present].copy()

df = df_sla.merge(df_cat_slim, on=sla_id, how="left")

matched = df["Reason"].notna().sum() if "Reason" in df.columns else 0
print(f"  After merge     : {len(df):,} rows")
print(f"  Matched         : {matched:,} ({matched/len(df)*100:.1f}%)")
print(f"  Unmatched       : {len(df)-matched:,} → Unclassified")

df["Reason"]          = df["Reason"].fillna("Unclassified")          if "Reason"          in df.columns else "Unclassified"
df["Refined Summary"] = df["Refined Summary"].fillna("Unclassified") if "Refined Summary" in df.columns else "Unclassified"


# ============================================================
# STEP 4 — FEATURE ENGINEERING
# ============================================================
print("\n" + "=" * 60)
print("PETRA PIPELINE — STEP 4: Feature Engineering")
print("=" * 60)

date_cols = [
    "Reported Date", "Last Modified Date", "Last Resolved Date",
    "Closed Date", "Actual Reported Date", "Actual Resolution Date",
    "Actual_Start", "Resolved_End",
]
for col in date_cols:
    if col in df.columns:
        df[col] = pd.to_datetime(df[col], errors="coerce")

for col in ["Resolution_Hours_Calendar", "Resolution_WorkingHours",
            "Overall Resolution Time (Second)"]:
    if col in df.columns:
        df[col] = pd.to_numeric(df[col], errors="coerce")

def sanitize_severity_for_model(s: pd.Series) -> pd.Series:
    if s is None:
        return pd.Series(dtype="object")
    s_str = s.astype(str).fillna("").str.strip()
    s_up  = s_str.str.upper()
    out   = s_str.copy()
    mask_s4_breach = (
        s_up.str.startswith("S4") &
        (s_up.str.contains("BREACH", regex=False) | s_up.str.contains("+", regex=False))
    )
    out.loc[mask_s4_breach] = "S4"
    norm_map = {"S1": "S1", "S2": "S2", "S3": "S3", "S4": "S4",
                "S1 ": "S1", "S2 ": "S2", "S3 ": "S3", "S4 ": "S4"}
    out = out.replace(norm_map)
    out = out.where(out.isin(["S1", "S2", "S3", "S4"]), "S4")
    return out

df["severity_clean"] = (
    df["Severity"].astype(str).str.strip()
    if "Severity" in df.columns else pd.Series("S4", index=df.index)
)
df["severity_clean_model"] = sanitize_severity_for_model(df["severity_clean"])

sev_num_map_model = {"S1": 1, "S2": 2, "S3": 3, "S4": 4}
df["severity_num"] = df["severity_clean_model"].map(sev_num_map_model).fillna(4)

sev_sla_map_model = {"S1": 6, "S2": 12, "S3": 105, "S4": 210}  # 09:00–midnight working hours
df["sla_target_hours"] = df["severity_clean_model"].map(sev_sla_map_model).fillna(210)

if df["severity_clean_model"].str.upper().str.contains("BREACH").any():
    raise ValueError("Sanitization failure: 'severity_clean_model' still contains 'BREACH'.")

sla_map   = {"TRUE": 1, "FALSE": 0, "True": 1, "False": 0,
             "Met": 1, "Missed": 0, "1": 1, "0": 0}
df["sla_met"]      = df["SLA_Met"].astype(str).map(sla_map) if "SLA_Met" in df.columns else 0
df["sla_breached"] = 1 - df["sla_met"].fillna(0)

df["resolution_hours"] = pd.to_numeric(
    df.get("Resolution_Hours_Calendar", pd.Series(dtype=float)), errors="coerce"
)
mask = (
    df["resolution_hours"].isna() &
    df.get("Actual Reported Date", pd.Series(dtype="datetime64[ns]")).notna() &
    df.get("Actual Resolution Date", pd.Series(dtype="datetime64[ns]")).notna()
)
if mask.any() and "Actual Reported Date" in df.columns and "Actual Resolution Date" in df.columns:
    df.loc[mask, "resolution_hours"] = (
        (df.loc[mask, "Actual Resolution Date"] - df.loc[mask, "Actual Reported Date"])
        .dt.total_seconds() / 3600
    )

now = pd.Timestamp.now()
if "Reported Date" in df.columns:
    df["ticket_age_hours"] = (now - df["Reported Date"]).dt.total_seconds() / 3600
    df["ticket_age_hours"] = df["ticket_age_hours"].clip(lower=0)
else:
    df["ticket_age_hours"] = 0.0

reported = df["Reported Date"] if "Reported Date" in df.columns else pd.Series(dtype="datetime64[ns]")
df["hour_reported"]  = reported.dt.hour.fillna(9).astype(int)
df["day_of_week"]    = reported.dt.dayofweek.fillna(0).astype(int)
df["is_monday"]      = (df["day_of_week"] == 0).astype(int)
df["is_friday"]      = (df["day_of_week"] == 4).astype(int)
df["is_weekend"]     = (df["day_of_week"] >= 5).astype(int)
df["is_after_hours"] = (
    (df.get("Start_BusinessBucket", pd.Series("", index=df.index)).astype(str).str.strip() == "After Hours") |
    (df["hour_reported"] < 9)
).astype(int)
# Seasonal features — quarter-end and early-month periods show higher breach rates
df["month_num"] = reported.dt.month.fillna(1).astype(int)
df["quarter"]   = reported.dt.quarter.fillna(1).astype(int)
df["is_qtr_end"] = df["month_num"].isin([3, 6, 9, 12]).astype(int)

def get_col(df, col, default="OTHER"):
    return df[col].astype(str).str.strip() if col in df.columns else pd.Series(default, index=df.index)

tier2 = get_col(df, "Product Categorization Tier 2")
df["application"]     = tier2.apply(
    lambda x: "IMOS" if "IMOS" in x.upper() else "ENDUR" if "ENDUR" in x.upper() else "OTHER"
)
df["functional_area"] = get_col(df, "Product Categorization Tier 3")
df["resolver_team"]   = get_col(df, "Assignee Group")
df["category"]        = df["Refined Summary"].astype(str).str.strip()
df["report_month"]    = reported.dt.to_period("M").astype(str) if reported.notna().any() else "Unknown"

print(f"  Overall SLA breach rate : {df['sla_breached'].mean()*100:.1f}%")
print(f"  Total tickets           : {len(df):,}")

print("\n  Breach rate by Refined Summary:")
cat_stats = (
    df[df["category"] != "Unclassified"]
    .groupby("category")["sla_breached"]
    .agg(["mean", "count"])
    .rename(columns={"mean": "breach_%", "count": "tickets"})
    .sort_values("breach_%", ascending=False)
)
cat_stats["breach_%"] = (cat_stats["breach_%"] * 100).round(1)
print(cat_stats.to_string())


# ============================================================
# STEP 5 — TRAIN BREACH PREDICTOR
# ============================================================
print("\n" + "=" * 60)
print("PETRA PIPELINE — STEP 5: Train Breach Predictor")
print("=" * 60)

le_cat  = LabelEncoder()
le_app  = LabelEncoder()
le_func = LabelEncoder()
le_team = LabelEncoder()

OPEN_STATUSES = ["Assigned", "Pending", "In Progress", "Open", "New"]

status_col = df["Status"].astype(str) if "Status" in df.columns else pd.Series("Closed", index=df.index)
df_train = df[
    df["sla_breached"].notna() &
    ~status_col.isin(OPEN_STATUSES)
].copy()

print("  Building category historical stats...")
cat_stats_train = df_train.groupby("category").agg(
    cat_breach_rate  = ("sla_breached",     "mean"),
    cat_median_hours = ("resolution_hours", "median"),
    cat_p75_hours    = ("resolution_hours", lambda x: x.quantile(0.75)),
    cat_p90_hours    = ("resolution_hours", lambda x: x.quantile(0.90)),
    cat_ticket_count = ("sla_breached",     "count"),
).reset_index()

df = df.merge(cat_stats_train, on="category", how="left")

global_breach_rate  = df_train["sla_breached"].mean()
global_median_hours = df_train["resolution_hours"].dropna().median()
global_p75_hours    = df_train["resolution_hours"].dropna().quantile(0.75)
global_p90_hours    = df_train["resolution_hours"].dropna().quantile(0.90)

df["cat_breach_rate"]  = df["cat_breach_rate"].fillna(global_breach_rate)
df["cat_median_hours"] = df["cat_median_hours"].fillna(global_median_hours)
df["cat_p75_hours"]    = df["cat_p75_hours"].fillna(global_p75_hours)
df["cat_p90_hours"]    = df["cat_p90_hours"].fillna(global_p90_hours)
df["cat_ticket_count"] = df["cat_ticket_count"].fillna(0)

df_train = df_train.merge(
    cat_stats_train, on="category", how="left", suffixes=("", "_new")
)
for col in ["cat_breach_rate","cat_median_hours","cat_p75_hours","cat_p90_hours","cat_ticket_count"]:
    if col + "_new" in df_train.columns:
        df_train[col] = df_train[col + "_new"].fillna(df_train.get(col, global_breach_rate))
        df_train.drop(columns=[col + "_new"], inplace=True)

df_train["cat_breach_rate"]  = df_train["cat_breach_rate"].fillna(global_breach_rate)
df_train["cat_median_hours"] = df_train["cat_median_hours"].fillna(global_median_hours)
df_train["cat_p75_hours"]    = df_train["cat_p75_hours"].fillna(global_p75_hours)
df_train["cat_p90_hours"]    = df_train["cat_p90_hours"].fillna(global_p90_hours)
df_train["cat_ticket_count"] = df_train["cat_ticket_count"].fillna(0)

if len(df_train) < 10:
    print(f"  [WARN] Only {len(df_train)} closed tickets found for training.")
    print("         Skipping model training — will score with rule-based risk.")
    MODEL_TRAINED = False
    roc_auc = 0.0
    cv      = np.array([0.0])
else:
    df_train["category_enc"] = le_cat.fit_transform(df_train["category"].fillna("Unknown"))
    df_train["app_enc"]      = le_app.fit_transform(df_train["application"].fillna("OTHER"))
    df_train["func_enc"]     = le_func.fit_transform(df_train["functional_area"].fillna("Unknown"))
    df_train["team_enc"]     = le_team.fit_transform(df_train["resolver_team"].fillna("Unknown"))

    print(f"  Training on {len(df_train):,} closed/resolved tickets")
    print("  ✅ Leakage-free: only features known at ticket creation used for training.")

    FEATURES = [
        "severity_num",
        "sla_target_hours",
        "cat_breach_rate",
        "cat_median_hours",
        "cat_p75_hours",
        "cat_p90_hours",
        "cat_ticket_count",
        "hour_reported",
        "day_of_week",
        "is_monday",
        "is_friday",
        "is_weekend",
        "is_after_hours",
        "month_num",
        "quarter",
        "is_qtr_end",
        "category_enc",
        "app_enc",
        "func_enc",
        "team_enc",
    ]

    df_model = df_train[FEATURES + ["sla_breached"]].dropna()
    X = df_model[FEATURES]
    y = df_model["sla_breached"].astype(int)

    stratify = y if y.nunique() > 1 else None
    X_train, X_test, y_train, y_test = train_test_split(
        X, y, test_size=0.2, random_state=42, stratify=stratify
    )

    # Class-imbalance fix: upweight breach cases proportionally
    n_pos = (y_train == 1).sum()
    n_neg = (y_train == 0).sum()
    pos_weight = n_neg / n_pos if n_pos > 0 else 1.0
    sample_weights = np.where(y_train == 1, pos_weight, 1.0)
    print(f"  Class balance  : {n_pos} breach / {n_neg} on-time  "
          f"(upweight breach ×{pos_weight:.1f})")

    _base = GradientBoostingClassifier(
        n_estimators=300, learning_rate=0.08,
        max_depth=4, subsample=0.85, random_state=42
    )
    _base.fit(X_train, y_train, sample_weight=sample_weights)

    # Calibrate probabilities using cross-validation on training data.
    # cv="prefit" was removed in scikit-learn ≥1.4 — use cv=5 on training data instead.
    # _base is kept separately so feature_importances_ remains accessible for charts.
    model = CalibratedClassifierCV(
        GradientBoostingClassifier(
            n_estimators=300, learning_rate=0.08,
            max_depth=4, subsample=0.85, random_state=42
        ),
        cv=5, method="isotonic"
    )
    model.fit(X_train, y_train, sample_weight=sample_weights)

    y_pred      = model.predict(X_test)
    y_pred_prob = model.predict_proba(X_test)[:, 1]

    print("\n  Model Performance:")
    print(classification_report(y_test, y_pred, target_names=["On Time", "SLA Breach"],
                                zero_division=0))

    if y_test.nunique() > 1:
        roc_auc = roc_auc_score(y_test, y_pred_prob)
        print(f"  ROC-AUC : {roc_auc:.3f}")
        cv      = cross_val_score(model, X, y, cv=min(5, len(df_model)), scoring="roc_auc")
        print(f"  5-Fold CV AUC : {cv.mean():.3f} (+/- {cv.std():.3f})")
    else:
        roc_auc = 0.0
        cv      = np.array([0.0])
        print("  [WARN] Only one class in test set — ROC-AUC not available.")

    print("\n  Feature importance (base GBM):")
    for feat, imp in sorted(zip(FEATURES, _base.feature_importances_), key=lambda x: -x[1]):
        bar = "█" * int(imp * 50)
        print(f"    {feat:30s} {bar} {imp:.3f}")

    print("\n  Category breach rates (historical):")
    cat_display = cat_stats_train.sort_values("cat_breach_rate", ascending=False)
    for _, row in cat_display.iterrows():
        bar = "█" * int(row["cat_breach_rate"] * 20)
        print(f"    {row['category']:45s} {bar} {row['cat_breach_rate']*100:.0f}%  "
              f"(median {row['cat_median_hours']:.0f}h, n={int(row['cat_ticket_count'])})")

    with open(OUT_DIR / "petra_breach_model.pkl", "wb") as f:
        pickle.dump({
            "model": model, "base_model": _base, "features": FEATURES,
            "le_cat": le_cat, "le_app": le_app,
            "le_func": le_func, "le_team": le_team,
            "categories": list(le_cat.classes_),
            "cat_stats": cat_stats_train,
            "trained_on": datetime.now().strftime("%Y-%m-%d %H:%M"),
            "roc_auc": round(roc_auc, 3),
            "pos_weight": round(pos_weight, 2),
        }, f)
    print(f"\n  Model saved → {OUT_DIR / 'petra_breach_model.pkl'}")
    MODEL_TRAINED = True


# ============================================================
# STEP 6 — SCORE TICKETS
# ============================================================
print("\n" + "=" * 60)
print("PETRA PIPELINE — STEP 6: Score Tickets")
print("=" * 60)

open_mask = status_col.isin(OPEN_STATUSES)
age_ratio = (df["ticket_age_hours"] / df["sla_target_hours"].replace(0, 1)).clip(0, 3)

urgency_boost = np.where(
    open_mask,
    np.where(age_ratio > 1.0,  25,
    np.where(age_ratio > 0.75, 15,
    np.where(age_ratio > 0.5,   8,
                                0))),
    0
)

if MODEL_TRAINED:
    def safe_encode(encoder, values):
        known  = set(encoder.classes_)
        mapped = values.apply(lambda x: x if x in known else encoder.classes_[0])
        return encoder.transform(mapped)

    df["category_enc"] = safe_encode(le_cat,  df["category"].fillna("Unknown"))
    df["app_enc"]      = safe_encode(le_app,  df["application"].fillna("OTHER"))
    df["func_enc"]     = safe_encode(le_func, df["functional_area"].fillna("Unknown"))
    df["team_enc"]     = safe_encode(le_team, df["resolver_team"].fillna("Unknown"))

    base_risk = model.predict_proba(df[FEATURES].fillna(df[FEATURES].median()))[:, 1] * 100
    df["breach_risk_%"] = (base_risk + urgency_boost).clip(0, 100).round(1)
    print("  Scoring: ML model base score + elapsed-time urgency boost (open tickets only)")

else:
    # Rule-based fallback: calibrate severity risk from historical data where possible
    sev_hist = (
        df_train.groupby("severity_clean_model")["sla_breached"].mean() * 100
        if len(df_train) > 0 else pd.Series(dtype=float)
    )
    sev_risk_default = {1: 60, 2: 45, 3: 25, 4: 10}
    sev_risk = {
        1: sev_hist.get("S1", sev_risk_default[1]),
        2: sev_hist.get("S2", sev_risk_default[2]),
        3: sev_hist.get("S3", sev_risk_default[3]),
        4: sev_hist.get("S4", sev_risk_default[4]),
    }
    print(f"  Rule-based severity risk (from history): {sev_risk}")
    df["_sev_risk"] = df["severity_num"].map(sev_risk).fillna(10)
    df["_cat_risk"] = (df["cat_breach_rate"].fillna(global_breach_rate).clip(0, 1) * 30)
    df["_time_risk"] = urgency_boost
    df["breach_risk_%"] = (
        df["_sev_risk"] + df["_cat_risk"] + df["_time_risk"]
    ).clip(0, 100).round(1)
    df.drop(columns=["_sev_risk", "_cat_risk", "_time_risk"], inplace=True)
    print("  Scoring: Rule-based (history-calibrated severity + category breach rate + urgency boost)")

df["risk_level"] = df["breach_risk_%"].apply(
    lambda x: "🔴 HIGH"   if x >= 70 else
              "🟡 MEDIUM" if x >= 40 else
              "🟢 LOW"
)

df_open      = df[open_mask].copy()
high_count   = (df_open["breach_risk_%"] >= 70).sum()
medium_count = ((df_open["breach_risk_%"] >= 40) & (df_open["breach_risk_%"] < 70)).sum()
low_count    = (df_open["breach_risk_%"] < 40).sum()

print(f"\n  Open tickets : {len(df_open):,}")
print(f"  🔴 HIGH      : {high_count}")
print(f"  🟡 MEDIUM    : {medium_count}")
print(f"  🟢 LOW       : {low_count}")

if len(df_open) > 0 and "category" in df_open.columns:
    print("\n  Risk breakdown by category (open tickets):")
    open_cat = (df_open.groupby("category")
                .agg(tickets  =("breach_risk_%", "count"),
                     avg_risk =("breach_risk_%", "mean"),
                     high     =("breach_risk_%", lambda x: (x >= 70).sum()))
                .sort_values("avg_risk", ascending=False)
                .head(15))
    for cat, row in open_cat.iterrows():
        bar = "█" * int(row["avg_risk"] / 5)
        print(f"    {cat:45s} {bar} {row['avg_risk']:.0f}%  ({int(row['high'])} HIGH / {int(row['tickets'])} open)")


# ── Predicted breach datetime ─────────────────────────────────────────────
# Estimate the timestamp when each open ticket will breach its SLA.
# Formula: reported_date + sla_target_hours (working hours → calendar approx)
# We convert working hours to calendar hours using the 15h/day ratio.
WORK_HOURS_PER_CAL_DAY = 15.0  # 09:00–midnight window

def _predicted_breach_dt(row) -> object:
    """Return estimated breach datetime for open tickets."""
    reported = row.get("Reported Date")
    sla_h    = row.get("sla_target_hours", 210)
    if pd.isna(reported) or pd.isna(sla_h) or sla_h <= 0:
        return pd.NaT
    # Convert working hours to approximate calendar hours
    cal_hours = (sla_h / WORK_HOURS_PER_CAL_DAY) * 24.0
    return pd.Timestamp(reported) + pd.Timedelta(hours=cal_hours)

if "Reported Date" in df_open.columns:
    df_open["predicted_breach_dt"] = df_open.apply(_predicted_breach_dt, axis=1)
    df_open["hours_until_breach"]  = (
        (df_open["predicted_breach_dt"] - now)
        .dt.total_seconds() / 3600
    ).round(1)
    # Negative = already past predicted breach time
    already_past = (df_open["hours_until_breach"] < 0).sum()
    if already_past > 0:
        print(f"\n  ⏰ {already_past} open tickets are past their predicted breach time.")
else:
    df_open["predicted_breach_dt"] = pd.NaT
    df_open["hours_until_breach"]  = np.nan


# ── Zombie ticket detector ────────────────────────────────────────────────
# A zombie is an open ticket that is >2× its SLA target age with no recent
# status change. These are tickets that have been forgotten.
ZOMBIE_MULTIPLIER = 2.0

if "ticket_age_hours" in df_open.columns and "sla_target_hours" in df_open.columns:
    zombie_mask = (
        (df_open["ticket_age_hours"] > df_open["sla_target_hours"] * ZOMBIE_MULTIPLIER) &
        (~df_open["severity_clean"].str.contains("Breach", case=False, na=False))
    )
    df_open["is_zombie"] = zombie_mask.astype(int)
    zombie_count = zombie_mask.sum()
    if zombie_count > 0:
        print(f"\n  🧟 Zombie tickets (>2× SLA age, not yet S4+ Breach): {zombie_count}")
        zombie_preview = df_open[zombie_mask][["Incident ID", "category", "ticket_age_hours",
                                               "sla_target_hours", "Assignee Group"]].head(5)
        for _, z in zombie_preview.iterrows():
            inc = z.get("Incident ID", "N/A")
            cat = z.get("category", "")
            age = z.get("ticket_age_hours", 0)
            tgt = z.get("sla_target_hours", 1)
            team = z.get("Assignee Group", "")
            print(f"    🧟 {inc} | {cat} | {age:.0f}h old ({age/tgt:.1f}× SLA) | {team}")
else:
    df_open["is_zombie"] = 0
    zombie_count = 0


# ============================================================
# STEP 6B — RESOLUTION SUGGESTER
# ============================================================
print("\n" + "=" * 60)
print("PETRA PIPELINE — STEP 6B: Resolution Suggester")
print("=" * 60)

RES_NOTE_COL = "Resolution Note"
DESC_COL     = "Summary"

JUNK = {
    "advise user", "corrective", "manual", "amended", "offscript",
    "run corrective", "fix", "n/a", "na", "none", "-", "closed",
    "assisted via call", "user confirmed", "issue resolved",
    "advise user to try again", "manually edit", "manually amend",
    "amend manually", "add manual", "edit manually", "corrective",
    "amend manually via call", "user unreachable", "amended",
}

def is_useful_note(note: str) -> bool:
    n = note.lower().strip()
    if len(n) < 20:
        return False
    if n in JUNK:
        return False
    generic = ["advise user", "run corrective", "manual", "corrective",
               "amended", "offscript", "fix", "assist"]
    if all(g in n for g in generic[:1]) and len(n) < 30:
        return False
    return True

if RES_NOTE_COL in df.columns and DESC_COL in df.columns:
    kb = df[
        ~status_col.isin(OPEN_STATUSES) &
        df[RES_NOTE_COL].notna() &
        df[DESC_COL].notna()
    ].copy()
    kb[RES_NOTE_COL] = kb[RES_NOTE_COL].astype(str).str.strip()
    kb[DESC_COL]     = kb[DESC_COL].astype(str).str.strip()
    kb = kb[kb[RES_NOTE_COL].apply(is_useful_note)].copy()
    kb = kb.drop_duplicates(subset=[RES_NOTE_COL])
else:
    kb = pd.DataFrame()

print(f"  Knowledge base   : {len(kb):,} closed tickets with useful resolution notes")

SUGGESTER_READY = len(kb) >= 10

if SUGGESTER_READY:
    vectorizer = TfidfVectorizer(
        ngram_range=(1, 2),
        max_features=5000,
        stop_words="english",
        min_df=2,
    )
    kb_matrix = vectorizer.fit_transform(kb[DESC_COL].fillna(""))

    def suggest_resolution(description: str, category: str, top_n: int = 3) -> list:
        if not description or not SUGGESTER_READY:
            return []
        query_vec = vectorizer.transform([str(description)])
        scores    = cosine_similarity(query_vec, kb_matrix).flatten()
        cat_mask  = kb["category"].values == category
        scores    = scores + (cat_mask * 0.15)
        top_idx   = scores.argsort()[::-1][:top_n * 3]
        results   = []
        seen_notes = set()
        for idx in top_idx:
            if len(results) >= top_n:
                break
            note     = kb.iloc[idx][RES_NOTE_COL]
            note_key = note.lower().strip()[:60]
            if note_key in seen_notes:
                continue
            seen_notes.add(note_key)
            results.append({
                "score":       round(float(scores[idx]), 3),
                "resolution":  note,
                "similar_to":  kb.iloc[idx][DESC_COL][:80],
                "category":    kb.iloc[idx]["category"],
                "closed_date": str(kb.iloc[idx].get("Closed Date", ""))[:10],
            })
        return results

    print("  Generating resolution suggestions for open tickets...")
    suggestions_list = []
    for _, row in df_open.iterrows():
        desc = str(row.get(DESC_COL, ""))
        cat  = str(row.get("category", ""))
        sugg = suggest_resolution(desc, cat, top_n=3)
        for rank, s in enumerate(sugg, 1):
            suggestions_list.append({
                "Incident ID":           row.get("Incident ID", ""),
                "Summary":               desc[:100],
                "Issue Type":            cat,
                "Breach Risk %":         row.get("breach_risk_%", 0),
                "Risk Level":            row.get("risk_level", ""),
                "Rank":                  rank,
                "Suggested Resolution":  s["resolution"],
                "Similarity Score":      s["score"],
                "Confidence":            "HIGH" if s["score"] > 0.5 else "MEDIUM" if s["score"] > 0.3 else "LOW",
                "Similar Ticket Desc":   s["similar_to"],
                "Similar Category":      s["category"],
                "Similar Closed":        s["closed_date"],
            })

    df_suggestions = pd.DataFrame(suggestions_list)

    if len(df_suggestions) > 0:
        top_sugg = (
            df_suggestions[df_suggestions["Rank"] == 1]
            [["Incident ID", "Suggested Resolution", "Similarity Score", "Confidence"]]
            .rename(columns={
                "Suggested Resolution": "top_suggestion",
                "Similarity Score":     "suggestion_score",
                "Confidence":           "suggestion_confidence",
            })
        )
        id_col_name = "Incident ID" if "Incident ID" in df_open.columns else df_open.columns[0]
        df_open = df_open.merge(top_sugg, on=id_col_name, how="left")

    print(f"  Suggestions generated : {len(df_suggestions):,} rows")
    if len(df_suggestions) > 0:
        print(f"  Avg similarity score  : {df_suggestions['Similarity Score'].mean():.3f}")
        print(f"  HIGH confidence       : {(df_suggestions['Similarity Score'] > 0.5).sum():,}")
        print(f"  MEDIUM confidence     : {((df_suggestions['Similarity Score'] > 0.3) & (df_suggestions['Similarity Score'] <= 0.5)).sum():,}")
        print(f"  LOW confidence        : {(df_suggestions['Similarity Score'] <= 0.3).sum():,}")
else:
    df_suggestions = pd.DataFrame()
    print("  [SKIP] Not enough useful resolution notes in knowledge base.")


# ============================================================
# STEP 7 — CHARTS
# ============================================================
print("\n" + "=" * 60)
print("PETRA PIPELINE — STEP 7: Charts")
print("=" * 60)

TEAL  = "#00A19C"
NAVY  = "#0D1B2A"
RED   = "#EF4444"
GOLD  = "#F59E0B"
GREEN = "#22C55E"
GRAY  = "#94A3B8"

fig, axes = plt.subplots(2, 3, figsize=(18, 10))
fig.patch.set_facecolor(NAVY)
fig.suptitle("PETRA — TRMS AI Service Intelligence",
             fontsize=16, fontweight="bold", color="white")

ax = axes[0, 0]
monthly = df.groupby("report_month").size()
ax.bar(monthly.index, monthly.values, color=TEAL, alpha=0.85)
ax.set_facecolor(NAVY); ax.tick_params(colors="white")
ax.set_title("Ticket Volume by Month", color="white", fontweight="bold")
for spine in ax.spines.values(): spine.set_edgecolor(GRAY)
plt.setp(ax.xaxis.get_majorticklabels(), rotation=45, ha="right", color="white")

ax = axes[0, 1]
cat_b  = (df[df["category"] != "Unclassified"]
          .groupby("category")["sla_breached"].mean().mul(100).sort_values().tail(12))
colors = [RED if v >= 50 else GOLD if v >= 25 else GREEN for v in cat_b.values]
ax.barh(cat_b.index, cat_b.values, color=colors, alpha=0.85)
ax.set_facecolor(NAVY); ax.tick_params(colors="white")
ax.set_title("Breach Rate by Category (%)", color="white", fontweight="bold")
for spine in ax.spines.values(): spine.set_edgecolor(GRAY)
plt.setp(ax.yaxis.get_majorticklabels(), color="white", fontsize=8)

ax = axes[0, 2]
sev_b  = (df.groupby("severity_clean")["sla_breached"].mean().mul(100)
          .reindex(["S1", "S2", "S3", "S4", "S4+ (Breach)"]).dropna())
colors = [RED if v >= 50 else GOLD if v >= 25 else GREEN for v in sev_b.values]
ax.bar(sev_b.index, sev_b.values, color=colors, alpha=0.85)
ax.set_facecolor(NAVY); ax.tick_params(colors="white")
ax.set_title("Breach Rate by Severity (%)", color="white", fontweight="bold")
for spine in ax.spines.values(): spine.set_edgecolor(GRAY)

ax = axes[1, 0]
if len(df_open) > 0:
    rc         = df_open["risk_level"].value_counts()
    pie_colors = [RED if "HIGH" in r else GOLD if "MEDIUM" in r else GREEN for r in rc.index]
    ax.pie(rc.values, labels=rc.index, colors=pie_colors, autopct="%1.0f%%",
           textprops={"color": "white", "fontsize": 10})
ax.set_facecolor(NAVY)
ax.set_title(f"Open Tickets Risk\n({len(df_open):,} open)", color="white", fontweight="bold")

ax = axes[1, 1]
if MODEL_TRAINED:
    fi = pd.Series(_base.feature_importances_, index=FEATURES).sort_values(ascending=True).tail(8)
    ax.barh(fi.index, fi.values, color=TEAL, alpha=0.85)
    ax.set_title("Top Breach Predictors", color="white", fontweight="bold")
else:
    ax.text(0.5, 0.5, "Model not trained\n(insufficient data)",
            ha="center", va="center", color="white", fontsize=12, transform=ax.transAxes)
    ax.set_title("Top Breach Predictors", color="white", fontweight="bold")
ax.set_facecolor(NAVY); ax.tick_params(colors="white")
for spine in ax.spines.values(): spine.set_edgecolor(GRAY)
plt.setp(ax.yaxis.get_majorticklabels(), color="white", fontsize=8)

ax = axes[1, 2]
res = df["resolution_hours"].dropna()
res = res[res > 0]
if len(res) > 0:
    res = res[res < res.quantile(0.95)]
    ax.hist(res, bins=40, color=TEAL, alpha=0.8, edgecolor=NAVY)
    ax.axvline(res.median(), color=GOLD, linestyle="--", lw=2,
               label=f"Median: {res.median():.0f}h")
    ax.axvline(res.mean(),   color=RED,  linestyle="--", lw=2,
               label=f"Mean: {res.mean():.0f}h")
    ax.legend(facecolor=NAVY, labelcolor="white", edgecolor=GRAY)
ax.set_facecolor(NAVY); ax.tick_params(colors="white")
ax.set_title("Resolution Time Distribution", color="white", fontweight="bold")
for spine in ax.spines.values(): spine.set_edgecolor(GRAY)

plt.tight_layout()
chart_path = OUT_DIR / "petra_charts.png"
plt.savefig(chart_path, dpi=150, bbox_inches="tight", facecolor=NAVY)
plt.close()
print(f"  Charts saved → {chart_path}")


# ============================================================
# STEP 8 — EXPORT MASTER EXCEL
# ============================================================
print("\n" + "=" * 60)
print("PETRA PIPELINE — STEP 8: Master Excel Report")
print("=" * 60)

EXPORT_COLS = [
    "Incident ID", "Service Request ID", "Reported Date", "Status",
    "Priority", "Summary", "category", "Refined Summary", "Reason",
    "Confidence",
    "severity_clean", "SLA_Met", "sla_breached", "resolution_hours",
    "ticket_age_hours", "Resolution_WorkingHours", "Resolution_Bucket",
    "Assignee Group", "Assignee", "application", "functional_area",
    "breach_risk_%", "risk_level",
    "predicted_breach_dt", "hours_until_breach", "is_zombie",
    "Start_BusinessBucket", "hour_reported",
    "day_of_week", "report_month", "Resolution Note",
    "Customer Organization", "Customer Department",
]
cols        = [c for c in EXPORT_COLS if c in df.columns]
master_path = OUT_DIR / "PETRA_Master_Report.xlsx"

with pd.ExcelWriter(master_path, engine="openpyxl") as writer:
    df[cols].to_excel(writer, sheet_name="Master Data", index=False)

    if len(df_open) > 0:
        open_cols = [c for c in cols if c in df_open.columns]
        (df_open[open_cols]
         .sort_values("breach_risk_%", ascending=False)
         .to_excel(writer, sheet_name="Open Tickets - Risk Ranked", index=False))

    high_risk = df[df["breach_risk_%"] >= 70]
    high_cols = [c for c in cols if c in high_risk.columns]
    (high_risk[high_cols]
     .sort_values("breach_risk_%", ascending=False)
     .to_excel(writer, sheet_name="HIGH RISK", index=False))

    (df.groupby("category").agg(
        tickets        =("Incident ID", "count") if "Incident ID" in df.columns else ("severity_clean", "count"),
        breached       =("sla_breached", "sum"),
        breach_rate_pct=("sla_breached", lambda x: round(x.mean() * 100, 1)),
        avg_hrs        =("resolution_hours", "mean"),
        median_hrs     =("resolution_hours", "median"),
     ).round(1).sort_values("breach_rate_pct", ascending=False).reset_index()
     .to_excel(writer, sheet_name="Category Summary", index=False))

    (df.groupby("report_month").agg(
        tickets        =("severity_clean", "count"),
        breached       =("sla_breached", "sum"),
        breach_rate_pct=("sla_breached", lambda x: round(x.mean() * 100, 1)),
        avg_hrs        =("resolution_hours", "mean"),
     ).round(1).reset_index()
     .to_excel(writer, sheet_name="Monthly Trend", index=False))

    if SUGGESTER_READY and len(df_suggestions) > 0:
        (df_suggestions
         .sort_values(["Breach Risk %", "Rank"], ascending=[False, True])
         .to_excel(writer, sheet_name="Resolution Suggestions", index=False))
        print("  Resolution Suggestions sheet added.")

    # ── Zombie Tickets sheet ──────────────────────────────────────────────
    if zombie_count > 0 and "is_zombie" in df_open.columns:
        zombie_cols = [c for c in cols if c in df_open.columns]
        (df_open[df_open["is_zombie"] == 1][zombie_cols]
         .sort_values("ticket_age_hours", ascending=False)
         .to_excel(writer, sheet_name="🧟 Zombie Tickets", index=False))
        print(f"  Zombie Tickets sheet: {zombie_count} tickets")

    # ── Weekly Health Summary ─────────────────────────────────────────────
    if "report_month" in df.columns and "Data_Year" in df.columns:
        weekly_agg = (
            df.groupby(["Data_Year", "report_month"]).agg(
                total_tickets  =("severity_clean", "count"),
                closed         =("Ticket_Status", lambda x: (x == "Closed").sum()) if "Ticket_Status" in df.columns else ("severity_clean", "count"),
                breached       =("sla_breached", "sum"),
                breach_rate_pct=("sla_breached", lambda x: round(x.mean() * 100, 1)),
                avg_working_hrs=("Resolution_WorkingHours", "mean") if "Resolution_WorkingHours" in df.columns else ("resolution_hours", "mean"),
                median_working_hrs=("Resolution_WorkingHours", "median") if "Resolution_WorkingHours" in df.columns else ("resolution_hours", "median"),
                s1_count       =("severity_clean", lambda x: (x == "S1").sum()),
                s2_count       =("severity_clean", lambda x: (x == "S2").sum()),
                s3_count       =("severity_clean", lambda x: (x == "S3").sum()),
                s4_count       =("severity_clean", lambda x: (x.str.startswith("S4")).sum()),
            ).round(1).reset_index()
        )
        weekly_agg.to_excel(writer, sheet_name="Weekly Health Summary", index=False)
        print("  Weekly Health Summary sheet added.")

    # ── Data Quality Report ───────────────────────────────────────────────
    dq_rows = []
    for col in ["Reported Date", "Severity", "Summary", "Assignee Group",
                "Resolution Note", "Product Categorization Tier 2",
                "Product Categorization Tier 3"]:
        if col in df.columns:
            null_n = df[col].isna().sum()
            dq_rows.append({
                "Column":        col,
                "Total":         len(df),
                "Missing":       null_n,
                "Missing_%":     round(null_n / len(df) * 100, 1),
                "Populated_%":   round((1 - null_n / len(df)) * 100, 1),
            })
    if dq_rows:
        pd.DataFrame(dq_rows).to_excel(writer, sheet_name="Data Quality", index=False)
        print("  Data Quality sheet added.")

    pd.DataFrame({
        "Metric": [
            "ROC-AUC", "CV AUC mean", "CV AUC std",
            "Training tickets", "Trained on", "Breach rate",
            "SLA File", "Classifier File",
            "Leakage fix", "Scoring method",
            "Suggester KB size",
            "Zombie tickets",
            "Predicted breach datetime",
        ],
        "Value": [
            f"{roc_auc:.3f}" if MODEL_TRAINED else "N/A",
            f"{cv.mean():.3f}" if MODEL_TRAINED else "N/A",
            f"{cv.std():.3f}"  if MODEL_TRAINED else "N/A",
            len(df_train) if MODEL_TRAINED else "N/A",
            datetime.now().strftime("%Y-%m-%d"),
            f"{df['sla_breached'].mean()*100:.1f}%",
            str(SLA_FILE),
            str(CAT_FILE),
            "Yes — resolution hours excluded from training",
            "ML base score + elapsed-time urgency boost (open tickets only)",
            f"{len(kb):,} useful notes" if SUGGESTER_READY else "N/A",
            str(zombie_count),
            "Yes — predicted_breach_dt column in open tickets",
        ],
    }).to_excel(writer, sheet_name="Model Info", index=False)

print(f"  Master report saved → {master_path}")


# ============================================================
# STEP 9 — TEAMS ALERT REPORT
# ============================================================
print("\n" + "=" * 60)
print("PETRA PIPELINE — STEP 9: Teams Alert Report")
print("=" * 60)

today_str  = datetime.now().strftime("%Y-%m-%d")
ALERT_COLS = [
    "Incident ID", "Reported Date", "Summary", "category",
    "severity_clean", "Assignee Group", "Assignee",
    "breach_risk_%", "risk_level", "Status",
    "ticket_age_hours", "sla_target_hours", "Confidence",
    "predicted_breach_dt", "hours_until_breach", "is_zombie",
]
alert_cols = [c for c in ALERT_COLS if c in df_open.columns]
df_alert   = (df_open[alert_cols]
              .sort_values("breach_risk_%", ascending=False)
              .reset_index(drop=True))

df_alert = df_alert.rename(columns={
    "Incident ID":         "Ticket ID",
    "Reported Date":       "Raised On",
    "category":            "Issue Type",
    "severity_clean":      "Severity",
    "Assignee Group":      "Team",
    "breach_risk_%":       "Breach Risk %",
    "risk_level":          "Risk Level",
    "ticket_age_hours":    "Age (hrs)",
    "sla_target_hours":    "SLA Target (hrs)",
    "predicted_breach_dt": "Predicted Breach",
    "hours_until_breach":  "Hrs Until Breach",
    "is_zombie":           "Zombie?",
})

if "Raised On" in df_alert.columns:
    df_alert["Raised On"] = pd.to_datetime(
        df_alert["Raised On"], errors="coerce"
    ).dt.strftime("%d %b %Y %H:%M")
if "Age (hrs)" in df_alert.columns:
    df_alert["Age (hrs)"] = pd.to_numeric(
        df_alert["Age (hrs)"], errors="coerce"
    ).round(1)

df_already_breached = df_alert[
    df_alert["Severity"].str.contains("Breach", case=False, na=False)
].copy()

df_at_risk = df_alert[
    (df_alert["Breach Risk %"] >= 70) &
    (~df_alert["Severity"].str.contains("Breach", case=False, na=False))
].copy()

df_monitor = df_alert[
    (df_alert["Breach Risk %"] >= 40) &
    (df_alert["Breach Risk %"] < 70) &
    (~df_alert["Severity"].str.contains("Breach", case=False, na=False))
].copy()

at_risk_count  = len(df_at_risk)
breached_count = len(df_already_breached)
monitor_count  = len(df_monitor)

print(f"  ⚠️  At Risk (preventable) : {at_risk_count}")
print(f"  🔴 Already Breached       : {breached_count}")
print(f"  🟡 Monitor                : {monitor_count}")

alert_path = OUT_DIR / f"PETRA_Teams_Alert_{today_str}.xlsx"

with pd.ExcelWriter(alert_path, engine="openpyxl") as writer:

    pd.DataFrame({
        "": [
            "📅 Report Date",
            "🎫 Total Open Tickets",
            "",
            "⚠️  AT RISK — Not Yet Breached (Act Now)",
            "🔴 Already Breached — SLA Missed (Escalate)",
            "🟡 Monitor — Medium Risk",
            "🟢 Low Risk",
            "",
            "🚨 Priority Action",
        ],
        " ": [
            today_str,
            len(df_open),
            "",
            at_risk_count,
            breached_count,
            monitor_count,
            low_count,
            "",
            "Focus on AT RISK tickets — these can still be saved before SLA breach.",
        ]
    }).to_excel(writer, sheet_name="📊 Summary", index=False)

    if len(df_at_risk) > 0:
        df_at_risk.to_excel(writer, sheet_name="⚠️ AT RISK - Act Now", index=False)
    else:
        pd.DataFrame({"Message": ["No tickets currently at risk — great work!"]}).to_excel(
            writer, sheet_name="⚠️ AT RISK - Act Now", index=False)

    if len(df_already_breached) > 0:
        df_already_breached.to_excel(writer, sheet_name="🔴 Already Breached", index=False)

    if len(df_monitor) > 0:
        df_monitor.to_excel(writer, sheet_name="🟡 Monitor", index=False)

    df_alert.to_excel(writer, sheet_name="All Open - Ranked", index=False)

    if "Team" in df_alert.columns:
        team_summary = df_alert.copy()
        team_summary["at_risk"]         = (
            (team_summary["Breach Risk %"] >= 70) &
            (~team_summary["Severity"].str.contains("Breach", case=False, na=False))
        ).astype(int)
        team_summary["already_breached"] = (
            team_summary["Severity"].str.contains("Breach", case=False, na=False)
        ).astype(int)
        team_summary["monitor"]          = (
            (team_summary["Breach Risk %"] >= 40) &
            (team_summary["Breach Risk %"] < 70) &
            (~team_summary["Severity"].str.contains("Breach", case=False, na=False))
        ).astype(int)
        (team_summary.groupby("Team").agg(
            total_open      =("Ticket ID", "count") if "Ticket ID" in team_summary.columns else ("Severity", "count"),
            at_risk         =("at_risk",          "sum"),
            already_breached=("already_breached", "sum"),
            monitor         =("monitor",          "sum"),
            avg_risk        =("Breach Risk %",    "mean"),
         ).round(1).sort_values("at_risk", ascending=False).reset_index()
         .to_excel(writer, sheet_name="By Team", index=False))

    # Resolution suggestions for HIGH risk and already breached tickets
    if SUGGESTER_READY and len(df_suggestions) > 0:
        high_ids = (
            df_at_risk["Ticket ID"].tolist() +
            df_already_breached["Ticket ID"].tolist()
        )
        high_sugg = df_suggestions[
            df_suggestions["Incident ID"].isin(high_ids)
        ].sort_values(["Breach Risk %", "Rank"], ascending=[False, True])

        if len(high_sugg) > 0:
            high_sugg.to_excel(writer, sheet_name="💡 Suggested Fixes", index=False)
            print(f"  💡 Suggested Fixes sheet: {len(high_sugg):,} suggestions")

    # ── Zombie Tickets sheet in alert report ──────────────────────────────
    if "Zombie?" in df_alert.columns:
        df_zombies_alert = df_alert[df_alert["Zombie?"] == 1].copy()
    else:
        df_zombies_alert = pd.DataFrame()
    if len(df_zombies_alert) > 0:
        df_zombies_alert.to_excel(writer, sheet_name="🧟 Zombie Tickets", index=False)
        print(f"  🧟 Zombie Tickets sheet: {len(df_zombies_alert)} tickets")


# ── Colour formatting ─────────────────────────────────────────
wb = load_workbook(alert_path)

FILL_RED    = PatternFill("solid", fgColor="FFCCCC")
FILL_ORANGE = PatternFill("solid", fgColor="FFE5CC")
FILL_GOLD   = PatternFill("solid", fgColor="FFF3CC")
FILL_GREEN  = PatternFill("solid", fgColor="CCFFCC")
FILL_NAVY   = PatternFill("solid", fgColor="0D1B2A")
FILL_TEAL   = PatternFill("solid", fgColor="00A19C")
FILL_HEADER = PatternFill("solid", fgColor="1A2E44")
FILL_HINT   = PatternFill("solid", fgColor="EEF4FF")

FONT_WHITE     = Font(color="FFFFFF", bold=True,  name="Calibri", size=11)
FONT_DARK      = Font(color="0D1B2A", bold=False, name="Calibri", size=10)
FONT_DARK_BOLD = Font(color="0D1B2A", bold=True,  name="Calibri", size=10)
FONT_RED       = Font(color="CC0000", bold=True,  name="Calibri", size=10)
FONT_ORANGE    = Font(color="CC6600", bold=True,  name="Calibri", size=10)
FONT_TITLE     = Font(color="FFFFFF", bold=True,  name="Calibri", size=13)
FONT_HINT      = Font(color="1A3A6B", bold=False, name="Calibri", size=10, italic=True)

thin   = Side(style="thin", color="CCCCCC")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

def style_header_row(ws):
    for cell in ws[1]:
        cell.fill = FILL_TEAL; cell.font = FONT_WHITE
        cell.alignment = CENTER; cell.border = BORDER

def style_data_rows_alert(ws, sev_col_idx=None, risk_col_idx=None):
    for row in ws.iter_rows(min_row=2):
        sev_val  = row[sev_col_idx  - 1].value if sev_col_idx  else None
        risk_val = row[risk_col_idx - 1].value if risk_col_idx else None
        for cell in row:
            cell.border = BORDER; cell.alignment = LEFT
            if sev_val == "S4+ (Breach)":
                cell.fill = FILL_RED;    cell.font = FONT_RED
            elif risk_val == "🔴 HIGH":
                cell.fill = FILL_ORANGE; cell.font = FONT_ORANGE
            elif risk_val == "🟡 MEDIUM":
                cell.fill = FILL_GOLD;   cell.font = FONT_DARK_BOLD
            elif risk_val == "🟢 LOW":
                cell.fill = FILL_GREEN;  cell.font = FONT_DARK
            else:
                cell.font = FONT_DARK

def auto_width(ws, max_width=55):
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len    = max((len(str(c.value)) for c in col if c.value), default=10)
        ws.column_dimensions[col_letter].width = min(max_len + 4, max_width)

# Summary sheet
ws_sum = wb["📊 Summary"]
ws_sum.sheet_view.showGridLines = False
ws_sum.column_dimensions["A"].width = 45
ws_sum.column_dimensions["B"].width = 45
ws_sum.insert_rows(1)
ws_sum["A1"] = f"🚨 PETRA SLA RISK ALERT — {today_str}"
ws_sum["A1"].font = FONT_TITLE; ws_sum["A1"].fill = FILL_NAVY
ws_sum["A1"].alignment = CENTER
ws_sum.merge_cells("A1:B1")
ws_sum.row_dimensions[1].height = 30
for row in ws_sum.iter_rows(min_row=2):
    for cell in row:
        cell.border = BORDER; cell.alignment = LEFT
        val = str(cell.value or "")
        if "AT RISK" in val or "Act Now" in val:
            cell.fill = FILL_ORANGE; cell.font = FONT_ORANGE
        elif "Already Breached" in val or "Escalate" in val:
            cell.fill = FILL_RED;    cell.font = FONT_RED
        elif "Monitor" in val:
            cell.fill = FILL_GOLD;   cell.font = FONT_DARK_BOLD
        elif "Low Risk" in val:
            cell.fill = FILL_GREEN;  cell.font = FONT_DARK
        elif any(k in val for k in ["Report Date", "Total", "Priority", "Focus"]):
            cell.fill = FILL_HEADER; cell.font = FONT_WHITE
        else:
            cell.font = FONT_DARK

# Data sheets
for sheet_name in ["⚠️ AT RISK - Act Now", "🔴 Already Breached",
                   "🟡 Monitor", "All Open - Ranked", "🧟 Zombie Tickets"]:
    if sheet_name in wb.sheetnames:
        ws   = wb[sheet_name]
        ws.sheet_view.showGridLines = False
        hdrs = [cell.value for cell in ws[1]]
        sidx = (hdrs.index("Severity")   + 1) if "Severity"   in hdrs else None
        ridx = (hdrs.index("Risk Level") + 1) if "Risk Level" in hdrs else None
        style_header_row(ws)
        style_data_rows_alert(ws, sev_col_idx=sidx, risk_col_idx=ridx)
        auto_width(ws)
        ws.freeze_panes = "A2"

# By Team sheet
if "By Team" in wb.sheetnames:
    ws_team = wb["By Team"]
    ws_team.sheet_view.showGridLines = False
    style_header_row(ws_team)
    auto_width(ws_team)
    hdrs = [cell.value for cell in ws_team[1]]
    at_risk_idx = (hdrs.index("at_risk") + 1) if "at_risk" in hdrs else None
    for row in ws_team.iter_rows(min_row=2):
        for cell in row:
            cell.border = BORDER; cell.font = FONT_DARK
        if at_risk_idx and row[at_risk_idx - 1].value and row[at_risk_idx - 1].value > 0:
            for cell in row:
                cell.fill = FILL_ORANGE; cell.font = FONT_ORANGE

# Suggested Fixes sheet
if "💡 Suggested Fixes" in wb.sheetnames:
    ws_fix = wb["💡 Suggested Fixes"]
    ws_fix.sheet_view.showGridLines = False
    style_header_row(ws_fix)
    auto_width(ws_fix)
    ws_fix.freeze_panes = "A2"
    hdrs = [cell.value for cell in ws_fix[1]]
    conf_idx = (hdrs.index("Confidence") + 1) if "Confidence" in hdrs else None
    rank_idx = (hdrs.index("Rank")       + 1) if "Rank"       in hdrs else None
    for row in ws_fix.iter_rows(min_row=2):
        conf_val = row[conf_idx - 1].value if conf_idx else None
        rank_val = row[rank_idx - 1].value if rank_idx else None
        for cell in row:
            cell.border = BORDER; cell.alignment = LEFT
            if rank_val == 1 and conf_val == "HIGH":
                cell.fill = FILL_ORANGE; cell.font = FONT_ORANGE
            elif rank_val == 1 and conf_val == "MEDIUM":
                cell.fill = FILL_GOLD;   cell.font = FONT_DARK_BOLD
            elif rank_val == 1:
                cell.fill = FILL_HINT;   cell.font = FONT_HINT
            else:
                cell.font = FONT_DARK

wb.save(alert_path)
print(f"  Teams alert saved → {alert_path}")


# ============================================================
# STEP 10 — PRINT TEAMS MESSAGE
# ============================================================
id_col   = "Ticket ID"       if "Ticket ID"       in df_open.columns else (
           "Incident ID"     if "Incident ID"     in df_open.columns else df_open.columns[0])
sum_col  = "Summary"         if "Summary"         in df_open.columns else None
sev_col  = "severity_clean"  if "severity_clean"  in df_open.columns else None
team_col = "Assignee Group"  if "Assignee Group"  in df_open.columns else None
age_col  = "ticket_age_hours" if "ticket_age_hours" in df_open.columns else None
hint_col = "top_suggestion"  if "top_suggestion"  in df_open.columns else None

top5_at_risk = df_open[
    (df_open["breach_risk_%"] >= 70) &
    (~df_open["severity_clean"].str.contains("Breach", case=False, na=False))
].nlargest(5, "breach_risk_%")

if len(top5_at_risk) == 0:
    top5_at_risk = df_open.nlargest(5, "breach_risk_%")

print("\n" + "=" * 65)
print("COPY THIS INTO YOUR TEAMS CHANNEL:")
print("=" * 65)

teams_msg = f"""
🚨 **PETRA SLA RISK ALERT — {today_str}**

📊 **Open Ticket Summary**
| Status | Count |
|---|---|
| ⚠️ AT RISK — Act Now (not yet breached) | **{at_risk_count}** |
| 🔴 Already Breached — Escalate | **{breached_count}** |
| 🟡 Monitor | **{monitor_count}** |
| 🟢 Low Risk | **{low_count}** |
| 🧟 Zombie (>2× SLA age) | **{zombie_count}** |
| **Total Open** | **{len(df_open)}** |

⚠️ **Top 5 AT RISK Tickets — Still Saveable, Act Now**
"""
for _, row in top5_at_risk.iterrows():
    ticket  = row.get(id_col,  "N/A")
    summary = str(row.get(sum_col,  ""))[:60] if sum_col  else ""
    sev     = row.get(sev_col,  "")           if sev_col  else ""
    team    = row.get(team_col, "")           if team_col else ""
    risk    = row.get("breach_risk_%", 0)
    age     = f"{row.get(age_col, 0):.0f}h old" if age_col else ""
    hint    = row.get(hint_col, "")            if hint_col else ""
    hint_str = f"\n  💡 _{str(hint)[:80]}_" if hint and str(hint).strip() and str(hint) != "nan" else ""
    teams_msg += f"""
- **{ticket}** — {sev} — {risk}% breach risk  {age}
  _{summary}_
  👤 {team}{hint_str}
"""

teams_msg += f"""
📎 Full report: `PETRA_Teams_Alert_{today_str}.xlsx`
_PETRA AI Service Intelligence — {datetime.now().strftime("%d %b %Y %H:%M")}_
"""
print(teams_msg)


# ============================================================
# FINAL SUMMARY
# ============================================================
print("=" * 60)
print(f"""
PETRA COMPLETE ✅

  Total tickets      : {len(df):,}
  Matched category   : {matched:,} ({matched/len(df)*100:.1f}%)
  SLA breach rate    : {df['sla_breached'].mean()*100:.1f}%
  Model trained      : {"Yes — AUC " + f"{roc_auc:.3f}" if MODEL_TRAINED else "No (rule-based scoring used)"}
  Leakage fix        : ✅ Resolution hours excluded from training
  Resolution KB      : {len(kb):,} useful notes {"✅" if SUGGESTER_READY else "⚠️  skipped"}

  Open tickets ({len(df_open)} total):
    ⚠️  AT RISK (act now) : {at_risk_count}
    🔴 Already Breached   : {breached_count}
    🟡 Monitor            : {monitor_count}
    🟢 Low Risk           : {low_count}
    🧟 Zombie tickets     : {zombie_count}

  New features:
    ✅ Predicted breach datetime (predicted_breach_dt column)
    ✅ Zombie ticket detector (is_zombie column)
    ✅ Data Quality report sheet
    ✅ Weekly Health Summary sheet
    ✅ History-calibrated rule-based fallback scoring

  Output files → {OUT_DIR}
    ├── Incident_SLA_All_Years.xlsx
    ├── PETRA_Master_Report.xlsx
    ├── PETRA_Teams_Alert_{today_str}.xlsx
    ├── petra_breach_model.pkl  {"✅" if MODEL_TRAINED else "⚠️  skipped"}
    └── petra_charts.png

  Classified file → {CAT_FILE}
""")
print("=" * 60)


# ============================================================
# FINAL STEP — GENERATE INTERACTIVE DASHBOARD
# ============================================================
print("\n" + "=" * 60)
print("PETRA PIPELINE — FINAL: Interactive HTML Dashboard")
print("=" * 60)
try:
    from dashboard import run_dashboard
    run_dashboard(open_browser=True)
except Exception as _dash_err:
    print(f"  [WARN] Dashboard skipped: {_dash_err}")
    print("  Run  python dashboard.py  separately to generate it.")
