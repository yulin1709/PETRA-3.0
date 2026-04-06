# sla.py
# -*- coding: utf-8 -*-
"""
PETRA — SLA Computation (Multi-Year)
Processes all year sheets (2026, 2025, 2024 etc.) from the latest
Incident Raw Data file, computes SLA columns for each, saves combined
enriched file to PETRA OUTPUT folder.
"""

from __future__ import annotations
import os, re, shutil
from pathlib import Path
from datetime import datetime, time, timedelta

import numpy as np
import pandas as pd
from openpyxl import load_workbook


# ============================================================
# PATH RESOLUTION
# ============================================================
def _candidate_bases() -> list[Path]:
    bases: list[Path] = []
    user = Path(os.environ.get("USERPROFILE", ""))
    p_base = user / "PETRONAS"
    if p_base.exists():
        bases.append(p_base)
    for envvar in ("OneDriveCommercial", "OneDrive"):
        v = os.environ.get(envvar)
        if v and os.path.isdir(v):
            bases.append(Path(v))
    od_guess = user / "OneDrive - PETRONAS"
    if od_guess.exists():
        bases.append(od_guess)
    seen, uniq = set(), []
    for b in bases:
        s = str(b)
        if s not in seen:
            seen.add(s); uniq.append(b)
    return uniq

BASES = _candidate_bases()

def _first_existing(*rel_parts: str) -> Path | None:
    for b in BASES:
        p = b.joinpath(*rel_parts)
        if p.exists():
            return p
    return None

_inc_env = os.environ.get("INCIDENT_ROOT")
if _inc_env:
    INCIDENT_ROOT = Path(_inc_env)
else:
    INCIDENT_ROOT = _first_existing("TRMS Internal - myGenie+ Extract")
    if INCIDENT_ROOT is None:
        INCIDENT_ROOT = Path(os.environ["USERPROFILE"]) / r"PETRONAS\TRMS Internal - myGenie+ Extract"

_hol_env = os.environ.get("HOLIDAY_CSV_PATH")
if _hol_env:
    HOLIDAY_CSV_PATH = Path(_hol_env)
else:
    user_base = Path(os.environ.get("USERPROFILE", "")) / "PETRONAS"
    preferred_petco_path = (
        user_base
        / "PETCO Trading Digital - myGenie Ticket Analysis"
        / "Others Ref"
        / "Holiday.csv"
    )
    if preferred_petco_path.exists():
        HOLIDAY_CSV_PATH = preferred_petco_path
    else:
        candidate = _first_existing(
            "Service Management", "3_Service Level Management",
            "8_SLR_Reporting", "myGenie Ticket Analysis", "Others Ref", "Holiday.csv",
        )
        HOLIDAY_CSV_PATH = candidate if candidate else preferred_petco_path

OUT_DIR = Path(os.environ["USERPROFILE"]) / "OneDrive - PETRONAS" / "Desktop" / "PETRA OUTPUT"
OUT_DIR.mkdir(parents=True, exist_ok=True)


# ============================================================
# AUTO-SELECT LATEST FILE
# ============================================================
def find_latest_incident_file(folder: Path) -> Path:
    pattern = re.compile(
        r"^Incident Raw Data - (\d{1,2}) (\w+) (\d{4})\.xlsx$",
        re.IGNORECASE
    )
    candidates = []
    for f in folder.iterdir():
        if not f.is_file():
            continue
        m = pattern.match(f.name)
        if m:
            try:
                dt = pd.to_datetime(
                    f"{m.group(1)} {m.group(2)} {m.group(3)}", dayfirst=True
                )
                candidates.append((dt, f))
            except Exception:
                continue
    if not candidates:
        raise FileNotFoundError(
            f"No 'Incident Raw Data - <dd> <Mon> <yyyy>.xlsx' found in:\n  {folder}"
        )
    candidates.sort(key=lambda x: x[0], reverse=True)
    latest_dt, latest_file = candidates[0]
    print(f"  Selected input: {latest_file.name}  (date: {latest_dt.date()})")
    return latest_file


# ============================================================
# CONFIG
# ============================================================
# Working window: 09:00 to 00:00 (midnight) = 15 hours per working day.
# Weekend and Malaysian public holidays are excluded.
BUSINESS_START     = time(9, 0)
# BUSINESS_END is midnight — represented as 00:00 of the *next* calendar day
# in working_hours_between(). Do NOT use time(0,0) directly in comparisons.
BUSINESS_DAYS      = {0, 1, 2, 3, 4}   # Mon–Fri
WORK_HOURS_PER_DAY = 15               # 09:00–00:00 = 15 h

# SLA breach thresholds — working hours within the 09:00–midnight window.
# S1 =  6 working hours
# S2 = 12 working hours
# S3 =  7 working days × 15 h/day = 105 working hours
# S4 = 14 working days × 15 h/day = 210 working hours
SLA_THRESHOLDS = {
    "S1":   6,    # 6 working hours
    "S2":  12,    # 12 working hours
    "S3": 105,    # 7 working days  (09:00–midnight = 15 h/day)
    "S4": 210,    # 14 working days (09:00–midnight = 15 h/day)
}

DURATION_BUCKETS = [0, 4, 8, 12, 24, 48, 72, 105, 140, 210, 99999]
DURATION_LABELS  = ["<=4h","4-8h","8-12h","12-24h","24-48h","48-72h",
                    "72-105h","105-140h","140-210h",">210h"]

COL_REPORTED   = "Reported Date"
COL_SVT_TITLE  = "SVT Title"
END_CANDIDATES = [
    "Actual Resolution Date",
    "Closed Date",
    "Service Target Completed Date",
    "Last Resolved Date",
]

COMPUTED_COLS = [
    "Actual_Start", "Resolved_End", "Ticket_Status",
    "Resolution_Hours_Calendar", "Resolution_WorkingHours",
    "Severity", "SLA_Met",
    "Start_BusinessBucket", "Resolved_BusinessBucket",
    "Start_HourBucket", "Resolved_HourBucket",
    "Start_DayOfWeek", "Resolved_DayOfWeek",
    "Resolution_Bucket", "Month",
    "Data_Year",
]


# ============================================================
# HELPERS
# ============================================================
def to_datetime_any(v):
    if pd.isna(v):
        return pd.NaT
    if isinstance(v, (pd.Timestamp, datetime)):
        return pd.to_datetime(v)
    try:
        fv = float(v)
        if 20000 < fv < 60000:
            return pd.to_datetime(fv, unit="D", origin="1899-12-30")
    except Exception:
        pass
    try:
        return pd.to_datetime(str(v), errors="coerce")
    except Exception:
        return pd.NaT


def coalesce_cols(df: pd.DataFrame, candidates: list[str]) -> pd.Series:
    result = pd.Series([pd.NaT] * len(df), index=df.index)
    for c in candidates:
        if c in df.columns:
            mask = result.isna() & df[c].notna()
            result[mask] = df[c][mask]
    return result


def business_bucket(ts: pd.Timestamp) -> str:
    if pd.isna(ts):
        return ""
    if ts.weekday() not in BUSINESS_DAYS:
        return "Weekend"
    # Working window is 09:00–00:00 (midnight). Any time from 09:00 through
    # 23:59 on a weekday counts as Business Hours; 00:00–08:59 is After Hours.
    return "Business Hours" if ts.hour >= 9 else "After Hours"


def hour_bucket(ts: pd.Timestamp) -> str:
    if pd.isna(ts):
        return ""
    return f"{ts.hour:02d}:00-{(ts.hour + 1) % 24:02d}:00"


def working_hours_between(
    start: pd.Timestamp, end: pd.Timestamp, holidays: set
) -> float:
    """Count working hours between start and end.

    Working window per day: 09:00 – 00:00 (midnight) = 15 hours.
    Weekends and Malaysian public holidays (from Holiday.csv) are excluded.

    Midnight is represented as 00:00 of the *next* calendar day so that
    arithmetic stays correct (time(0,0) < time(9,0) in Python).
    """
    start = pd.Timestamp(start)
    end   = pd.Timestamp(end)
    if pd.isna(start) or pd.isna(end):
        return np.nan
    if end < start:
        # Data quality issue — log and return 0 rather than crashing
        return 0.0
    if end == start:
        return 0.0
    total = 0.0
    cur = start
    while cur.date() <= end.date():
        if cur.weekday() in BUSINESS_DAYS and cur.date().isoformat() not in holidays:
            # Working window: 09:00 on this day → 00:00 on the NEXT day
            ws = pd.Timestamp(datetime.combine(cur.date(), BUSINESS_START))
            we = pd.Timestamp(
                datetime.combine(cur.date() + timedelta(days=1), time(0, 0))
            )
            i_start = max(cur, ws)
            i_end   = min(end, we)
            if i_start < i_end:
                total += (i_end - i_start).total_seconds() / 3600.0
        cur = pd.Timestamp(
            datetime.combine((cur + pd.Timedelta(days=1)).date(), time(0, 0))
        )
    return total


def classify_severity(hours: float) -> str:
    if pd.isna(hours):
        return ""
    if hours <= SLA_THRESHOLDS["S1"]: return "S1"
    if hours <= SLA_THRESHOLDS["S2"]: return "S2"
    if hours <= SLA_THRESHOLDS["S3"]: return "S3"
    if hours <= SLA_THRESHOLDS["S4"]: return "S4"
    return "S4+ (Breach)"


def sla_met(row) -> object:
    if pd.isna(row["Resolution_WorkingHours"]) or row["Severity"] == "":
        return None
    return row["Severity"] != "S4+ (Breach)"


def safe_value(value):
    if value is None:
        return None
    if value is pd.NaT:
        return None
    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass
    if isinstance(value, float) and np.isnan(value):
        return None
    if hasattr(value, "item"):
        return value.item()
    if hasattr(value, '__class__') and value.__class__.__name__ == 'Period':
        return str(value)
    return value


# ============================================================
# DETECT YEAR SHEETS
# ============================================================
def detect_year_sheets(xl: pd.ExcelFile) -> list[tuple[int, str]]:
    year_sheets = []
    for sheet in xl.sheet_names:
        if re.fullmatch(r"20\d{2}", sheet.strip()):
            year_sheets.append((int(sheet.strip()), sheet))
        elif m := re.search(r"20\d{2}", sheet):
            year_sheets.append((int(m.group()), sheet))

    if not year_sheets:
        year_sheets = [(0, sheet) for sheet in xl.sheet_names]

    year_sheets.sort(key=lambda x: x[0], reverse=True)
    return year_sheets


# ============================================================
# PROCESS ONE SHEET
# ============================================================
def process_sheet(
    df_raw: pd.DataFrame,
    year_label: int,
    holidays: set,
    is_current_year: bool,
) -> pd.DataFrame:
    NOW = pd.Timestamp(datetime.now())
    df  = df_raw.copy()

    # Drop stale computed cols
    stale = [c for c in COMPUTED_COLS if c in df.columns]
    if stale:
        df.drop(columns=stale, inplace=True)

    # Parse dates
    if COL_REPORTED not in df.columns:
        print(f"    [WARN] '{COL_REPORTED}' not found in sheet — skipping.")
        return pd.DataFrame()

    df[COL_REPORTED] = df[COL_REPORTED].map(to_datetime_any)
    for c in END_CANDIDATES:
        if c in df.columns:
            df[c] = df[c].map(to_datetime_any)

    # Dedup — keep Resolution SVT rows if available
    inc_col = next(
        (c for c in df.columns if c.strip().lower() == "incident id"),
        next(
            (c for c in df.columns if "incident" in c.lower() and "id" in c.lower()),
            None
        )
    )

    if COL_SVT_TITLE in df.columns:
        res_mask = df[COL_SVT_TITLE].astype(str).str.contains(
            "Resolution", case=False, na=False
        )
        df_res = df[res_mask].copy()
        if not df_res.empty:
            df = df_res.copy()
    if inc_col:
        df = df.drop_duplicates(subset=[inc_col], keep="last").copy()

    total = len(df)
    if total == 0:
        return pd.DataFrame()

    # ── Data quality checks (run after Resolved_End is computed below) ───
    future_mask = df[COL_REPORTED] > NOW
    if future_mask.any():
        print(f"    [WARN] {future_mask.sum()} tickets have Reported Date in the future — check data.")

    # Tag year
    df["Data_Year"] = year_label if year_label > 0 else df[COL_REPORTED].dt.year.fillna(0).astype(int)

    # Resolved end — row-wise coalesce across all candidate columns
    df["Resolved_End"] = coalesce_cols(df, END_CANDIDATES).map(to_datetime_any)

    # Negative-duration check
    neg_mask = (
        df["Resolved_End"].notna() &
        df[COL_REPORTED].notna() &
        (df["Resolved_End"] < df[COL_REPORTED])
    )
    if neg_mask.any():
        print(f"    [WARN] {neg_mask.sum()} tickets have Resolved_End before Reported Date — "
              f"these will show 0 working hours.")

    # Ticket status — Open if no resolved end, Closed otherwise
    df["Ticket_Status"] = np.where(df["Resolved_End"].isna(), "Open", "Closed")

    # Effective end — always use NOW for open tickets
    effective_end = df["Resolved_End"].where(df["Resolved_End"].notna(), other=NOW)

    df["Actual_Start"] = df[COL_REPORTED]

    # Calendar hours (wall-clock elapsed)
    df["Resolution_Hours_Calendar"] = (
        (effective_end - df["Actual_Start"])
        .dt.total_seconds() / 3600.0
    )

    # ── Working hours: 09:00–00:00 on Mon–Fri, excl. public holidays ──────
    # This must be computed here because it is NOT present in the raw Excel.
    # effective_end is applied: for open tickets this is NOW (so working hours
    # keeps accumulating); for closed tickets it is the actual resolution date.
    print(f"    Computing Resolution_WorkingHours for {total:,} rows "
          f"(09:00–midnight window)…")
    df["Resolution_WorkingHours"] = [
        working_hours_between(s, e, holidays)
        for s, e in zip(df["Actual_Start"], effective_end)
    ]

    # Severity + SLA — based purely on working hours elapsed, open or closed
    df["Severity"] = df["Resolution_WorkingHours"].map(classify_severity)
    df["SLA_Met"]  = df.apply(sla_met, axis=1)

    # Dimension columns
    df["Start_BusinessBucket"]    = df["Actual_Start"].map(business_bucket)
    df["Resolved_BusinessBucket"] = df["Resolved_End"].map(business_bucket)
    df["Start_HourBucket"]        = df["Actual_Start"].map(hour_bucket)
    df["Resolved_HourBucket"]     = df["Resolved_End"].map(hour_bucket)
    df["Start_DayOfWeek"]         = df["Actual_Start"].dt.day_name()
    df["Resolved_DayOfWeek"]      = df["Resolved_End"].dt.day_name()
    df["Resolution_Bucket"]       = pd.cut(
        df["Resolution_WorkingHours"],
        bins=DURATION_BUCKETS, labels=DURATION_LABELS,
        include_lowest=True, right=True,
    ).astype(object)
    df["Month"] = df[COL_REPORTED].dt.to_period("M").astype(str)

    return df


# ============================================================
# MAIN FUNCTION
# ============================================================
def run_sla() -> Path:
    print("=" * 60)
    print("PETRA — SLA.py")
    print("=" * 60)
    print(f"  INCIDENT_ROOT : {INCIDENT_ROOT}")
    print(f"  Holiday CSV   : {HOLIDAY_CSV_PATH}")
    print(f"  Output dir    : {OUT_DIR}")

    if not INCIDENT_ROOT.exists():
        raise FileNotFoundError(f"INCIDENT_ROOT not found:\n  {INCIDENT_ROOT}")
    if not HOLIDAY_CSV_PATH.exists():
        raise FileNotFoundError(f"HOLIDAY_CSV_PATH not found:\n  {HOLIDAY_CSV_PATH}")

    # ── Holiday CSV — validate structure before parsing ──────────────────
    holiday_df = pd.read_csv(HOLIDAY_CSV_PATH)
    required_hol_cols = {"Year", "Month", "Day"}
    missing_hol_cols  = required_hol_cols - set(holiday_df.columns)
    if missing_hol_cols:
        raise ValueError(
            f"Holiday CSV is missing required columns: {missing_hol_cols}\n"
            f"  Found columns: {list(holiday_df.columns)}"
        )

    holiday_df["Month_parsed"] = pd.to_datetime(
        holiday_df["Month"].astype(str), format="%B", errors="coerce"
    ).dt.month
    mask = holiday_df["Month_parsed"].isna()
    holiday_df.loc[mask, "Month_parsed"] = pd.to_numeric(
        holiday_df.loc[mask, "Month"], errors="coerce"
    )

    # Validate Day and Year are numeric
    holiday_df["Day_num"]  = pd.to_numeric(holiday_df["Day"],  errors="coerce")
    holiday_df["Year_num"] = pd.to_numeric(holiday_df["Year"], errors="coerce")
    bad_rows = holiday_df["Day_num"].isna() | holiday_df["Year_num"].isna() | holiday_df["Month_parsed"].isna()
    if bad_rows.any():
        print(f"  [WARN] {bad_rows.sum()} holiday rows could not be parsed and will be skipped.")

    holiday_df = holiday_df[~bad_rows].copy()

    parsed_dates = pd.to_datetime(
        holiday_df["Year_num"].astype(int).astype(str) + "-" +
        holiday_df["Month_parsed"].astype(int).astype(str).str.zfill(2) + "-" +
        holiday_df["Day_num"].astype(int).astype(str).str.zfill(2),
        errors="coerce"
    ).dropna()

    HOLIDAYS: set[str] = set(parsed_dates.dt.date.astype(str))
    before_dedup = len(HOLIDAYS)
    # Dedup is implicit in set(), but log if source had duplicates
    raw_count = len(parsed_dates)
    if raw_count > before_dedup:
        print(f"  [INFO] {raw_count - before_dedup} duplicate holiday dates removed.")
    print(f"  Loaded {len(HOLIDAYS)} unique holiday dates.")

    SOURCE_FILE = find_latest_incident_file(INCIDENT_ROOT)
    DEST_FILE   = OUT_DIR / "Incident_SLA_All_Years.xlsx"

    try:
        shutil.copy2(SOURCE_FILE, DEST_FILE)
    except OSError:
        print("[WARN] shutil.copy2 failed — using openpyxl fallback.")
        wb_src = load_workbook(SOURCE_FILE, read_only=False, keep_vba=False)
        wb_src.save(DEST_FILE)
        wb_src.close()

    xl = pd.ExcelFile(DEST_FILE, engine="openpyxl")
    year_sheets = detect_year_sheets(xl)

    print(f"\n  Found {len(year_sheets)} year sheet(s):")
    for yr, sh in year_sheets:
        print(f"    {yr} → '{sh}'")

    current_year = datetime.now().year

    all_dfs: list[pd.DataFrame] = []
    year_summaries: list[dict]  = []

    for year_int, sheet_name in year_sheets:
        print(f"\n  {'=' * 50}")
        print(f"  Processing sheet: '{sheet_name}' (year label: {year_int})")
        print(f"  {'=' * 50}")

        try:
            df_raw = pd.read_excel(DEST_FILE, sheet_name=sheet_name, engine="openpyxl")
        except Exception as e:
            print(f"    [ERROR] Could not read sheet '{sheet_name}': {e}")
            continue

        print(f"    Raw rows: {len(df_raw):,}")

        if year_int == 0:
            if COL_REPORTED not in df_raw.columns:
                print(f"    [WARN] '{COL_REPORTED}' not found — skipping.")
                continue
            df_raw[COL_REPORTED] = df_raw[COL_REPORTED].map(to_datetime_any)
            df_raw["_parsed_year"] = df_raw[COL_REPORTED].dt.year.fillna(0).astype(int)

            unique_years = sorted(df_raw["_parsed_year"].unique(), reverse=True)
            unique_years = [y for y in unique_years if y > 2000]
            print(f"    Found {len(unique_years)} year(s) in Reported Date: {unique_years}")

            for yr in unique_years:
                df_yr_raw = df_raw[df_raw["_parsed_year"] == yr].copy()
                df_yr_raw.drop(columns=["_parsed_year"], inplace=True)
                print(f"\n    --- Year {yr}: {len(df_yr_raw):,} rows ---")

                is_current = (yr == datetime.now().year)
                df_proc    = process_sheet(df_yr_raw, yr, HOLIDAYS, is_current)

                if df_proc.empty:
                    print(f"    [WARN] No data after processing year {yr} — skipping.")
                    continue

                print(f"    Processed rows : {len(df_proc):,}")

                closed_df      = df_proc[df_proc["Ticket_Status"] == "Closed"]
                open_df        = df_proc[df_proc["Ticket_Status"] == "Open"]
                met_count      = df_proc["SLA_Met"].eq(True).sum()
                breached_count = df_proc["SLA_Met"].eq(False).sum()
                breach_rate    = (breached_count / len(df_proc) * 100) if len(df_proc) > 0 else 0

                print(f"    Closed={len(closed_df):,}  Open={len(open_df):,}  "
                      f"Met={met_count}  Breached={breached_count}  "
                      f"Breach rate={breach_rate:.1f}%")

                if "Severity" in df_proc.columns:
                    sev_counts = df_proc["Severity"].value_counts().sort_index()
                    print(f"    Severity: " + "  ".join(f"{s}={c}" for s, c in sev_counts.items()))

                year_summaries.append({
                    "Year":          yr,
                    "Sheet":         sheet_name,
                    "Total_Tickets": len(df_proc),
                    "Closed":        len(closed_df),
                    "Open":          len(open_df),
                    "SLA_Met":       int(met_count),
                    "SLA_Breached":  int(breached_count),
                    "Breach_Rate_%": round(breach_rate, 1),
                })
                all_dfs.append(df_proc)

        else:
            is_current = (year_int == datetime.now().year)
            df_proc    = process_sheet(df_raw, year_int, HOLIDAYS, is_current)

            if df_proc.empty:
                print(f"    [WARN] No data after processing — skipping.")
                continue

            print(f"    Processed rows : {len(df_proc):,}")

            closed_df      = df_proc[df_proc["Ticket_Status"] == "Closed"]
            open_df        = df_proc[df_proc["Ticket_Status"] == "Open"]
            met_count      = df_proc["SLA_Met"].eq(True).sum()
            breached_count = df_proc["SLA_Met"].eq(False).sum()
            breach_rate    = (breached_count / len(df_proc) * 100) if len(df_proc) > 0 else 0

            print(f"    Closed={len(closed_df):,}  Open={len(open_df):,}  "
                  f"Met={met_count}  Breached={breached_count}  "
                  f"Breach rate={breach_rate:.1f}%")

            if "Severity" in df_proc.columns:
                sev_counts = df_proc["Severity"].value_counts().sort_index()
                print(f"    Severity: " + "  ".join(f"{s}={c}" for s, c in sev_counts.items()))

            year_summaries.append({
                "Year":          year_int,
                "Sheet":         sheet_name,
                "Total_Tickets": len(df_proc),
                "Closed":        len(closed_df),
                "Open":          len(open_df),
                "SLA_Met":       int(met_count),
                "SLA_Breached":  int(breached_count),
                "Breach_Rate_%": round(breach_rate, 1),
            })
            all_dfs.append(df_proc)

    if not all_dfs:
        raise ValueError("No data processed from any sheet.")

    df_combined = pd.concat(all_dfs, ignore_index=True)
    print(f"\n  {'=' * 50}")
    print(f"  Combined total : {len(df_combined):,} rows across {len(all_dfs)} year(s)")

    print(f"\n  OVERALL SUMMARY BY YEAR:")
    print(f"  {'Year':<8} {'Total':>8} {'Closed':>8} {'Open':>6} {'Met':>6} {'Breached':>10} {'Breach%':>10}")
    print(f"  {'-'*60}")
    for s in year_summaries:
        print(f"  {str(s['Year']):<8} {s['Total_Tickets']:>8,} {s['Closed']:>8,} "
              f"{s['Open']:>6,} {s['SLA_Met']:>6,} {s['SLA_Breached']:>10,} "
              f"{s['Breach_Rate_%']:>9.1f}%")

    total_tickets  = sum(s["Total_Tickets"]  for s in year_summaries)
    total_closed   = sum(s["Closed"]         for s in year_summaries)
    total_open     = sum(s["Open"]           for s in year_summaries)
    total_met      = sum(s["SLA_Met"]        for s in year_summaries)
    total_breached = sum(s["SLA_Breached"]   for s in year_summaries)
    total_br       = (total_breached / total_tickets * 100) if total_tickets > 0 else 0
    print(f"  {'-'*60}")
    print(f"  {'TOTAL':<8} {total_tickets:>8,} {total_closed:>8,} "
          f"{total_open:>6,} {total_met:>6,} {total_breached:>10,} "
          f"{total_br:>9.1f}%")

    print(f"\n  Writing output → {DEST_FILE.name}")

    from openpyxl.styles import PatternFill, Font, Alignment
    FILL_HEADER = PatternFill("solid", fgColor="0D1B2A")
    FONT_WHITE  = Font(color="FFFFFF", bold=True, name="Calibri", size=10)
    CENTER      = Alignment(horizontal="center", vertical="center")

    wb_out = load_workbook(DEST_FILE)

    combined_sheet_name = "All Years (Combined)"
    if combined_sheet_name in wb_out.sheetnames:
        del wb_out[combined_sheet_name]
    ws_all = wb_out.create_sheet(title=combined_sheet_name, index=0)

    cols = list(df_combined.columns)
    for ci, col in enumerate(cols, 1):
        cell = ws_all.cell(row=1, column=ci, value=col)
        cell.fill = FILL_HEADER
        cell.font = FONT_WHITE
        cell.alignment = CENTER

    for ri, row in enumerate(df_combined.itertuples(index=False), 2):
        for ci, val in enumerate(row, 1):
            ws_all.cell(row=ri, column=ci, value=safe_value(val))

    summary_sheet_name = "Year Summary"
    if summary_sheet_name in wb_out.sheetnames:
        del wb_out[summary_sheet_name]
    ws_sum = wb_out.create_sheet(title=summary_sheet_name, index=1)

    sum_cols = ["Year", "Sheet", "Total_Tickets", "Closed", "Open",
                "SLA_Met", "SLA_Breached", "Breach_Rate_%"]
    for ci, col in enumerate(sum_cols, 1):
        cell = ws_sum.cell(row=1, column=ci, value=col)
        cell.fill = FILL_HEADER
        cell.font = FONT_WHITE
        cell.alignment = CENTER

    for ri, s in enumerate(year_summaries, 2):
        for ci, col in enumerate(sum_cols, 1):
            ws_sum.cell(row=ri, column=ci, value=s[col])

    ri = len(year_summaries) + 2
    ws_sum.cell(row=ri, column=1, value="TOTAL")
    ws_sum.cell(row=ri, column=3, value=total_tickets)
    ws_sum.cell(row=ri, column=4, value=total_closed)
    ws_sum.cell(row=ri, column=5, value=total_open)
    ws_sum.cell(row=ri, column=6, value=total_met)
    ws_sum.cell(row=ri, column=7, value=total_breached)
    ws_sum.cell(row=ri, column=8, value=round(total_br, 1))

    for year_int, sheet_name in year_sheets:
        matching = [d for d in all_dfs if not d.empty and
                    (d["Data_Year"] == year_int).any()]
        if not matching:
            continue
        df_yr = matching[0]

        if sheet_name in wb_out.sheetnames:
            del wb_out[sheet_name]
        ws_yr = wb_out.create_sheet(title=sheet_name)

        yr_cols = list(df_yr.columns)
        for ci, col in enumerate(yr_cols, 1):
            cell = ws_yr.cell(row=1, column=ci, value=col)
            cell.fill = FILL_HEADER
            cell.font = FONT_WHITE
            cell.alignment = CENTER

        for ri2, row in enumerate(df_yr.itertuples(index=False), 2):
            for ci, val in enumerate(row, 1):
                ws_yr.cell(row=ri2, column=ci, value=safe_value(val))

    wb_out.save(DEST_FILE)

    print(f"\n  ✅ SLA enrichment complete.")
    print(f"     {len(df_combined):,} rows written → {DEST_FILE}")
    print(f"     Sheets: '{combined_sheet_name}', '{summary_sheet_name}', "
          + ", ".join(f"'{sh}'" for _, sh in year_sheets))

    return DEST_FILE


if __name__ == "__main__":
    run_sla()